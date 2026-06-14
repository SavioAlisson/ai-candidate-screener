"""
Slack Intake Bot — Universal candidate submission to Ashby.

Runs as a Socket Mode app (no public URL needed). Supports:
  /intake      — single candidate modal form
  /intakebulk  — paste/CSV/XLSX multi-candidate preview flow (roles, resumes, dedup)

Usage:
  python3 slack_intake.py                    # run in foreground
  tmux new -s intake 'python3 slack_intake.py'  # persistent

Setup:
  1. Create Slack app at api.slack.com → enable Socket Mode
  2. Add /intake and /intakebulk slash commands
  3. Bot scopes: commands, chat:write, files:read, files:write, users:read,
                 reactions:read, reactions:write, channels:history, groups:history
  4. Event subscription: message.channels, message.groups, reaction_added
  5. Set env vars: SLACK_BOT_TOKEN (xoxb-), SLACK_APP_TOKEN (xapp-), ASHBY_API_KEY
  6. pip install openpyxl  (required for XLSX parsing in /intakebulk)

Env vars:
  SLACK_BOT_TOKEN  — Bot User OAuth Token (xoxb-)
  SLACK_APP_TOKEN  — App-Level Token for Socket Mode (xapp-)
  ASHBY_API_KEY    — Ashby API key (already in environment)
"""

from __future__ import annotations

import csv as _csv
import io as _io
import json
import logging
import os
import re
import threading
import time
import urllib.request
import uuid
from pathlib import Path
from html import escape as html_escape, unescape as html_unescape
from typing import Any, Dict, List, Optional, Set, Tuple

from slack_bolt import App
from slack_bolt.adapter.socket_mode import SocketModeHandler

logging.basicConfig(level=logging.INFO, format="%(asctime)s  %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

_DIR = Path(__file__).resolve().parent

# ── Purple Unicorn Hiring channel (broadcast target for all intake confirmations) ────
# Every /intake and /intakebulk submission posts a public confirmation here so the
# hiring team sees activity even if /intake was run in a DM.
HIRING_CHANNEL = "REPLACE_WITH_YOUR_HIRING_CHANNEL_ID"

# ── Candidate Labs Slack Connect channel ──────────────────────────
# External agency (Candidate Labs) posts LinkedIn URLs + PDF dossiers here.
# Posts from team_id KLARITY_TEAM_ID are internal (ignore); everything else is
# a Candidate Labs submission. Zero write-backs to this channel — verdict posts
# go to HIRING_CHANNEL so external parties don't see internal screening output.
CANDIDATE_LABS_CHANNEL = "REPLACE_WITH_YOUR_AGENCY_CHANNEL_ID"
KLARITY_TEAM_ID = "T1792L49E"
CANDIDATE_LABS_SOURCE = "Agencies: Candidate Labs"

# ── HubSpot Referral channel (refer_a_candidate) ─────────────────
# HubSpot bot posts "🎉 New referral" messages here whenever a Klarity employee
# submits a candidate via the HubSpot referral form. Each message has structured
# fields (Referrer, Candidate, Email, LinkedIn, Role, Strength, Notes). We parse,
# push to Ashby with source = HUBSPOT_REFERRAL_SOURCE, target the role's job via
# REFERRAL_ROLE_MAP (push_referrals.py), screen inline, and post the verdict to
# HIRING_CHANNEL. The refer_a_candidate channel itself stays silent (visible to
# referrers — no internal screening output leaks).
REFER_CANDIDATE_CHANNEL = os.environ.get("REFER_CANDIDATE_CHANNEL_ID", "")
HUBSPOT_REFERRAL_SOURCE = "Klarity Referral Campaign 2026"

# ── Source options (grouped, matching Ashby source tags exactly) ──
INTAKE_SOURCES = [
    {
        "label": "Sourced",
        "options": [
            ("Sourced: LinkedIn", "Sourced: LinkedIn"),
            ("Sourced: GitHub", "Sourced: GitHub"),
            ("Sourced: Juicebox", "Sourced: Juicebox"),
            ("Sourced: Sourcing Form", "Sourced: Sourcing Form"),
            ("Sourced: Recruiting Event", "Sourced: Recruiting Event"),
        ],
    },
    {
        "label": "Inbound",
        "options": [
            ("INBOUND: Dover Careers Page", "INBOUND: Dover Careers Page"),
        ],
    },
    {
        "label": "Referral",
        "options": [
            ("Referral: Referral", "Referral: Referral"),
        ],
    },
    {
        "label": "Agencies",
        "options": [
            ("Agencies: Hirewell", "Agencies: Hirewell"),
            ("Agencies: Candidate Labs", "Agencies: Candidate Labs"),
        ],
    },
]

# ── Load roles from .ashby_job_routing.json ──────────────────────

def _load_role_options() -> List[Dict]:
    """Load roles from routing config for the dropdown. Returns [{"text": display, "value": job_id}, ...]."""
    routing_file = _DIR / ".ashby_job_routing.json"
    if not routing_file.exists():
        return []
    data = json.loads(routing_file.read_text(encoding="utf-8"))
    options = []
    for role_name, info in data.get("roles", {}).items():
        if role_name == "Outbound Sourced":
            continue  # Added separately as "No target role"
        options.append({
            "text": info.get("job_title", role_name),
            "value": info["job_id"],
        })
    return options


OUTBOUND_JOB_ID = "REPLACE_WITH_YOUR_OUTBOUND_JOB_ID"
OUTBOUND_PROJECT_ID = "REPLACE_WITH_YOUR_OUTBOUND_PROJECT_ID"


# ── LinkedIn normalization (shared with push_to_ashby) ───────────

def _normalize_linkedin(url: str) -> str:
    if not url:
        return ""
    s = url.strip().lower()
    s = re.sub(r"https?://(www\.)?linkedin\.com", "", s, flags=re.I)
    s = s.rstrip("/")
    s = re.sub(r"\?.*$", "", s)
    return s


# ── Dedup helpers ────────────────────────────────────────────────

def _load_push_log() -> Dict[str, str]:
    path = _DIR / ".push_to_ashby_log.json"
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}


def _save_push_log(log: dict):
    path = _DIR / ".push_to_ashby_log.json"
    from ashby_bridge import write_json_atomic
    write_json_atomic(path, log, indent=2, ensure_ascii=False)


def _load_known_linkedins() -> set:
    """Load normalized LinkedIn URLs from screening_log.csv."""
    import csv
    known = set()
    log_path = _DIR / "screening_log.csv"
    if not log_path.exists():
        return known
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                li = (row.get("linkedin") or "").strip()
                if li:
                    known.add(_normalize_linkedin(li))
    except Exception:
        pass
    return known


def _ashby_candidate_url(candidate_id: str, application_id: str = "") -> str:
    """Build an Ashby candidate URL that actually resolves in Klarity's workspace.
    Plain /candidates/{id} 404s — the app expects a candidate-search context."""
    if not candidate_id:
        return ""
    if application_id:
        return f"https://app.ashbyhq.com/candidate-searches/new/right-side/candidates/{candidate_id}/applications/{application_id}/feed"
    return f"https://app.ashbyhq.com/candidate-searches/new/right-side/candidates/{candidate_id}"


def _lookup_first_application_id(candidate_id: str) -> str:
    """Fetch the first applicationId for a candidate via candidate.info. Empty string on failure."""
    if not candidate_id:
        return ""
    try:
        from ashby_bridge import _ashby_post
        r = _ashby_post("candidate.info", {"id": candidate_id})
        if r.get("success"):
            app_ids = (r.get("results") or {}).get("applicationIds") or []
            if app_ids:
                return app_ids[0]
    except Exception:
        pass
    return ""


def _check_duplicate(name: str, linkedin: str) -> Optional[Dict]:
    """Check if candidate already exists. Returns {"id": ..., "source": ...} or None."""
    li_norm = _normalize_linkedin(linkedin)

    # An empty normalized LinkedIn must NEVER match against the LinkedIn-keyed
    # caches — otherwise any candidate submitted without a LinkedIn URL collides
    # with every other empty-LinkedIn record. Fall through to the Ashby name
    # search (layer 3) which is the only safe identity check in that case.
    if li_norm:
        # Layer 1: push log
        push_log = _load_push_log()
        if li_norm in push_log:
            return {"id": push_log[li_norm], "source": "push log"}

        # Layer 2: screening log
        known = _load_known_linkedins()
        if li_norm in known:
            return {"id": "", "source": "screening log"}

    # Layer 3: Ashby name search. Ashby's candidate.search is fuzzy — searching
    # "Beau" returns "Elie Beaubrun" — so a hit only counts as a duplicate when
    # the returned name matches the searched name exactly (case/whitespace-
    # insensitive). False-match incident: Beau Davenport referral silently
    # dropped as a dup of Elie Beaubrun (2026-06-04).
    try:
        from ashby_bridge import search_candidate
        existing = search_candidate(name=name)
        if existing:
            def _norm_name(s: str) -> str:
                return re.sub(r"\s+", " ", (s or "").strip().lower())
            if _norm_name(existing.get("name", "")) == _norm_name(name):
                return {"id": existing.get("id", ""), "source": "Ashby"}
            logger.info(
                "Dedup layer-3: fuzzy Ashby hit %r ignored for search %r (names differ)",
                existing.get("name", ""), name,
            )
    except Exception:
        pass

    return None


# ── Ashby push ───────────────────────────────────────────────────

def _push_to_ashby(
    name: str,
    linkedin: str,
    email: str,
    source: str,
    job_id: str,
    notes_html: str,
) -> tuple[Optional[str], Optional[str]]:
    """Create candidate + application in Ashby. Returns (candidate_id, application_id) or (None, None)."""
    from ashby_bridge import create_application, add_note, _ashby_post, _resolve_source_id

    # Build payload — Ashby requires sourceId (UUID), not source (string)
    payload: Dict = {"name": name}
    source_id = _resolve_source_id(source)
    if source_id:
        payload["sourceId"] = source_id
    else:
        payload["source"] = source  # fallback for unknown sources
    if email:
        payload["emailAddresses"] = [{"value": email, "type": "Primary"}]

    # Create candidate
    result = _ashby_post("candidate.create", payload)
    if not result.get("success"):
        logger.error("candidate.create failed for %s: %s", name, result)
        return None, None
    candidate_id = result["results"]["id"]

    # Add social links via update (candidate.create ignores socialLinks)
    social_links = [{"type": "LinkedIn", "url": linkedin}]
    _ashby_post("candidate.update", {
        "candidateId": candidate_id,
        "socialLinks": social_links,
    })

    # Create application on target job
    app = create_application(candidate_id, job_id, source=source)
    application_id = (app or {}).get("id", "") or None

    # Ensure source sticks via application.changeSource (create alone isn't reliable)
    if application_id and source_id:
        _ashby_post("application.changeSource", {
            "applicationId": application_id,
            "sourceId": source_id,
        })

    # Add to Outbound Sourced project
    _ashby_post("candidate.addProject", {
        "candidateId": candidate_id,
        "projectId": OUTBOUND_PROJECT_ID,
    })

    # Add note
    if notes_html:
        add_note(candidate_id, notes_html)

    # Update push log — never write an empty key (would dedup-collide every
    # future no-LinkedIn referral against this one record).
    li_norm = _normalize_linkedin(linkedin)
    if li_norm:
        push_log = _load_push_log()
        push_log[li_norm] = candidate_id
        _save_push_log(push_log)

    logger.info("Created candidate %s → %s (job %s)", name, candidate_id[:12], job_id[:12])
    return candidate_id, application_id


# ── Resume upload to Ashby (presigned URL flow) ──────────────────

def _upload_resume_to_ashby(candidate_id: str, pdf_bytes: bytes, filename: str) -> bool:
    """Upload a PDF to Ashby's candidate resume field via presigned URL.

    Flow: file.createFileUploadHandle → PUT to presigned URL → candidate.uploadResume.
    Once uploaded, the resume shows in the CV section and is picked up by the
    screening pipeline on the next ascreen run (via resumeFileHandle).
    """
    from ashby_bridge import _ashby_post

    try:
        # 1. Get a presigned upload URL + file handle
        # Ashby requires filename, contentType, contentLength alongside fileUploadContext.
        safe_name = filename or f"{candidate_id[:8]}_resume.pdf"
        handle_resp = _ashby_post("file.createFileUploadHandle", {
            "fileUploadContext": "CandidateResume",
            "filename": safe_name,
            "contentType": "application/pdf",
            "contentLength": len(pdf_bytes),
        })
        if not handle_resp.get("success"):
            logger.error("createFileUploadHandle failed: %s", handle_resp)
            return False

        results = handle_resp.get("results", {})
        presigned_url = results.get("uploadUrl") or results.get("url", "")
        file_handle = results.get("fileHandle") or results.get("handle", "")
        fields = results.get("fields", {}) or {}
        if not presigned_url or not file_handle:
            logger.error("Missing uploadUrl or fileHandle in response: %s", results)
            return False

        # 2. Ashby returns an S3 presigned POST (multipart/form-data).
        # All returned fields go as form data; Content-Type must match the policy;
        # the file itself must be the LAST field.
        import requests
        form = dict(fields)
        form.setdefault("Content-Type", "application/pdf")
        files = {"file": (safe_name, pdf_bytes, "application/pdf")}
        post_resp = requests.post(presigned_url, data=form, files=files, timeout=60)
        if post_resp.status_code not in (200, 201, 204):
            logger.error("Presigned POST failed: %s %s", post_resp.status_code, post_resp.text[:500])
            return False

        # 3. Attach the uploaded file to the candidate as their resume
        attach_resp = _ashby_post("candidate.uploadResume", {
            "candidateId": candidate_id,
            "resumeHandle": file_handle,
        })
        if not attach_resp.get("success"):
            logger.error("candidate.uploadResume failed: %s", attach_resp)
            return False

        logger.info("Resume uploaded to Ashby for %s (%d bytes)", candidate_id[:12], len(pdf_bytes))
        return True

    except Exception as e:
        logger.error("Resume upload error for %s: %s", candidate_id[:12], e)
        return False


def _download_slack_file(url_private_download: str) -> Optional[bytes]:
    """Download a file from Slack using the bot token."""
    try:
        req = urllib.request.Request(
            url_private_download,
            headers={"Authorization": f"Bearer {os.environ['SLACK_BOT_TOKEN']}"},
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            return resp.read()
    except Exception as e:
        logger.error("Failed to download Slack file: %s", e)
        return None


def _notify_user(client, channel_id: str, user_id: str, text: str = "", blocks=None):
    """Best-effort private notification. Tries ephemeral in the invoking channel;
    on channel_not_found (DM/private channel bot isn't in), falls back to DMing the
    user directly via chat_postMessage with channel=user_id."""
    kwargs = {"text": text}
    if blocks is not None:
        kwargs["blocks"] = blocks
    if channel_id:
        try:
            client.chat_postEphemeral(channel=channel_id, user=user_id, **kwargs)
            return
        except Exception as e:
            logger.info("Ephemeral post failed (%s), falling back to DM", e)
    try:
        client.chat_postMessage(channel=user_id, **kwargs)
    except Exception as e:
        logger.error("DM fallback also failed for %s: %s", user_id, e)


# ── Modal builder ────────────────────────────────────────────────

def build_intake_modal(channel_id: str = "", default_source: str = "") -> dict:
    """Build the Block Kit modal view for /intake."""
    role_options = _load_role_options()

    # Role dropdown options
    role_opts = [
        {"text": {"type": "plain_text", "text": r["text"]}, "value": r["value"]}
        for r in role_options
    ]
    role_opts.append({
        "text": {"type": "plain_text", "text": "No target role / General Pool"},
        "value": OUTBOUND_JOB_ID,
    })

    # Source dropdown with option groups
    source_groups = []
    for group in INTAKE_SOURCES:
        opts = [
            {"text": {"type": "plain_text", "text": label}, "value": value}
            for label, value in group["options"]
        ]
        source_groups.append({
            "label": {"type": "plain_text", "text": group["label"]},
            "options": opts,
        })

    # Find default source option for initial_option
    source_initial = None
    if default_source:
        for group in INTAKE_SOURCES:
            for label, value in group["options"]:
                if value == default_source:
                    source_initial = {"text": {"type": "plain_text", "text": label}, "value": value}
                    break
            if source_initial:
                break

    # Build blocks
    blocks = [
        {
            "type": "input",
            "block_id": "name_block",
            "label": {"type": "plain_text", "text": "Candidate Name"},
            "element": {
                "type": "plain_text_input",
                "action_id": "name_input",
                "placeholder": {"type": "plain_text", "text": "Full name"},
            },
        },
        {
            "type": "input",
            "block_id": "linkedin_block",
            "label": {"type": "plain_text", "text": "LinkedIn URL"},
            "element": {
                "type": "url_text_input",
                "action_id": "linkedin_input",
                "placeholder": {"type": "plain_text", "text": "https://linkedin.com/in/..."},
            },
        },
        {
            "type": "input",
            "block_id": "role_block",
            "label": {"type": "plain_text", "text": "Target Role"},
            "element": {
                "type": "static_select",
                "action_id": "role_select",
                "placeholder": {"type": "plain_text", "text": "Select a role"},
                "options": role_opts,
            },
        },
    ]

    # Source dropdown
    source_element = {
        "type": "static_select",
        "action_id": "source_select",
        "placeholder": {"type": "plain_text", "text": "Select source"},
        "option_groups": source_groups,
    }
    if source_initial:
        source_element["initial_option"] = source_initial

    blocks.append({
        "type": "input",
        "block_id": "source_block",
        "label": {"type": "plain_text", "text": "Source"},
        "element": source_element,
    })

    # Optional fields
    blocks.extend([
        {
            "type": "input",
            "block_id": "email_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "Email"},
            "element": {
                "type": "email_text_input",
                "action_id": "email_input",
                "placeholder": {"type": "plain_text", "text": "candidate@example.com"},
            },
        },
        {
            "type": "input",
            "block_id": "referrer_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "Referred by"},
            "element": {
                "type": "plain_text_input",
                "action_id": "referrer_input",
                "placeholder": {"type": "plain_text", "text": "Name of referrer (for referral sources)"},
            },
        },
        {
            "type": "input",
            "block_id": "referrer_email_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "Referrer email"},
            "hint": {"type": "plain_text", "text": "Klarity email of the internal referrer — used to credit them in Ashby."},
            "element": {
                "type": "email_text_input",
                "action_id": "referrer_email_input",
                "placeholder": {"type": "plain_text", "text": "referrer@klarity.ai"},
            },
        },
        {
            "type": "input",
            "block_id": "notes_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "Notes"},
            "element": {
                "type": "plain_text_input",
                "action_id": "notes_input",
                "multiline": True,
                "placeholder": {"type": "plain_text", "text": "Any context about this candidate..."},
            },
        },
        {
            "type": "input",
            "block_id": "cv_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "CV / Resume Text (paste, fallback)"},
            "element": {
                "type": "plain_text_input",
                "action_id": "cv_input",
                "multiline": True,
                "placeholder": {"type": "plain_text", "text": "Paste resume text if you don't have a PDF"},
            },
        },
        {
            "type": "input",
            "block_id": "pdf_block",
            "optional": True,
            "label": {"type": "plain_text", "text": "Resume PDF"},
            "hint": {"type": "plain_text", "text": "Uploads directly to the candidate's CV section in Ashby."},
            "element": {
                "type": "file_input",
                "action_id": "pdf_input",
                "filetypes": ["pdf"],
                "max_files": 1,
            },
        },
    ])

    return {
        "type": "modal",
        "callback_id": "intake_modal",
        "title": {"type": "plain_text", "text": "Submit Candidate"},
        "submit": {"type": "plain_text", "text": "Submit"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": json.dumps({"channel_id": channel_id}),
        "blocks": blocks,
    }


# ── Slack event handlers ─────────────────────────────────────────

def handle_intake(ack, command, client):
    """Open the intake modal when /intake is invoked."""
    ack()
    channel_id = command.get("channel_id", "")
    modal = build_intake_modal(channel_id=channel_id, default_source="")
    client.views_open(trigger_id=command["trigger_id"], view=modal)


def handle_submission(ack, body, client, view):
    """Process the modal submission — create candidate in Ashby."""
    values = view["state"]["values"]
    metadata = json.loads(view.get("private_metadata", "{}"))
    channel_id = metadata.get("channel_id", "")
    user_id = body["user"]["id"]

    # Extract form values
    name = values["name_block"]["name_input"]["value"].strip()
    linkedin = values["linkedin_block"]["linkedin_input"]["value"].strip()
    role_selection = values["role_block"]["role_select"]["selected_option"]
    source_selection = values["source_block"]["source_select"]["selected_option"]

    role_name = role_selection["text"]["text"]
    job_id = role_selection["value"]
    source_tag = source_selection["value"]

    email = (values["email_block"]["email_input"].get("value") or "").strip()
    referrer = (values["referrer_block"]["referrer_input"].get("value") or "").strip()
    referrer_email = (values.get("referrer_email_block", {}).get("referrer_email_input", {}).get("value") or "").strip()
    notes = (values["notes_block"]["notes_input"].get("value") or "").strip()
    cv_text = (values["cv_block"]["cv_input"].get("value") or "").strip()

    # Pull uploaded PDF (if any) — file_input returns a list of file objects
    pdf_files = values.get("pdf_block", {}).get("pdf_input", {}).get("files") or []
    pdf_file = pdf_files[0] if pdf_files else None

    # Validate LinkedIn
    if "linkedin.com/in/" not in linkedin.lower():
        ack(response_action="errors", errors={
            "linkedin_block": "Must be a LinkedIn profile URL (linkedin.com/in/...)"
        })
        return

    ack()

    # Dedup check
    dup = _check_duplicate(name, linkedin)
    if dup:
        dup_app_id = _lookup_first_application_id(dup["id"]) if dup["id"] else ""
        ashby_link = _ashby_candidate_url(dup["id"], dup_app_id)
        msg = f":warning: *{name}* may already exist in Ashby (found via {dup['source']})."
        if ashby_link:
            msg += f"\n<{ashby_link}|View in Ashby>"
        msg += "\n\nIf this is a different person, contact the hiring team."
        _notify_user(client, channel_id, user_id, text=msg)
        return

    # Build note HTML
    note_parts = []
    note_parts.append(f"<b>Submitted via Slack by</b> <@{user_id}>")
    note_parts.append(f"<b>Source:</b> {source_tag}")
    if referrer:
        note_parts.append(f"<b>Referred by:</b> {referrer}")
    if notes:
        note_parts.append(f"<b>Notes:</b> {notes}")
    if cv_text:
        note_parts.append(f"<b>CV / Resume:</b><br>{cv_text[:5000]}")
    notes_html = "<br><br>".join(note_parts)

    # Push to Ashby
    candidate_id, application_id = _push_to_ashby(
        name=name,
        linkedin=linkedin,
        email=email,
        source=source_tag,
        job_id=job_id,
        notes_html=notes_html,
    )

    if not candidate_id:
        _notify_user(
            client, channel_id, user_id,
            text=f":x: Failed to create *{name}* in Ashby. Check logs.",
        )
        return

    # Credit the internal referrer in Ashby's "Credited To" field if we can
    # resolve them to a Klarity user. External referrers stay notes-only.
    if referrer or referrer_email:
        try:
            from ashby_bridge import (
                resolve_referrer_user_id, set_application_credited_to,
                set_referrer_fields,
            )
            # Always write Referrer Name + Referrer Email — campaign report needs both.
            set_referrer_fields(candidate_id, referrer_name=referrer, referrer_email=referrer_email)
            if application_id:
                ref_uid = resolve_referrer_user_id(email=referrer_email, name=referrer)
                if ref_uid:
                    set_application_credited_to(application_id, ref_uid)
        except Exception as e:
            logger.warning("Intake credit attribution failed for %s: %s", name, e)

    # If a PDF was uploaded in the modal, push it to Ashby's resume field
    resume_status = ""
    if pdf_file:
        file_name = pdf_file.get("name", "resume.pdf")
        download_url = pdf_file.get("url_private_download", "")
        if download_url:
            pdf_bytes = _download_slack_file(download_url)
            if pdf_bytes:
                uploaded = _upload_resume_to_ashby(candidate_id, pdf_bytes, file_name)
                resume_status = (
                    f":page_facing_up: Resume attached to Ashby CV section."
                    if uploaded else
                    f":warning: Candidate created, but resume upload failed — reply in thread with the PDF to retry."
                )
            else:
                resume_status = ":warning: Could not download the uploaded PDF from Slack."
        else:
            resume_status = ":warning: Uploaded file had no download URL."

    # Post confirmation to channel
    ashby_url = _ashby_candidate_url(candidate_id, application_id or "")
    confirm_blocks = [
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": (
                    f":white_check_mark: *Candidate submitted to Ashby*\n\n"
                    f"*Name:* {name}\n"
                    f"*LinkedIn:* <{linkedin}|Profile>\n"
                    f"*Target Role:* {role_name}\n"
                    f"*Source:* {source_tag}\n"
                    + (f"*Referred by:* {referrer}\n" if referrer else "")
                    + f"*Submitted by:* <@{user_id}>\n\n"
                    f"<{ashby_url}|View in Ashby>"
                ),
            },
        },
        {
            "type": "context",
            "elements": [
                {
                    "type": "mrkdwn",
                    "text": (
                        resume_status
                        or ":page_facing_up: _No PDF attached. Reply to this thread with a PDF resume to add one._"
                    ),
                },
            ],
        },
    ]

    # Always post the public confirmation to the Purple Unicorn Hiring channel so
    # the whole hiring team sees activity regardless of where /intake was run.
    confirmation_ts = ""
    try:
        result = client.chat_postMessage(
            channel=HIRING_CHANNEL,
            blocks=confirm_blocks,
            text=f"Candidate {name} submitted to Ashby",
            metadata={
                "event_type": "intake_candidate",
                "event_payload": {
                    "candidate_id": candidate_id,
                    "candidate_name": name,
                },
            },
        )
        confirmation_ts = result.get("ts", "")
        logger.info("Confirmation posted for %s in hiring channel (ts=%s)", name, confirmation_ts)
    except Exception as e:
        logger.warning("Hiring-channel post failed (%s), falling back to private notify", e)
        _notify_user(
            client, channel_id, user_id,
            text=f"Candidate {name} submitted to Ashby",
            blocks=confirm_blocks,
        )

    # Kick off inline screening in a background thread so the /intake view
    # submission returns immediately. The screening result is posted back in
    # the confirmation thread (or the hiring channel if ts is missing).
    _spawn_inline_screen(
        client=client,
        candidate_id=candidate_id,
        application_id=application_id or "",
        name=name,
        thread_ts=confirmation_ts,
    )


def handle_message(event, client):
    """Watch for PDF uploads in intake confirmation threads, plus dispatch
    Candidate Labs channel traffic to its dedicated auto-intake handler."""
    # Candidate Labs channel is handled with its own rules (auto-push, no chatter)
    if event.get("channel") == CANDIDATE_LABS_CHANNEL:
        try:
            handle_candidate_labs_message(event, client)
        except Exception as e:
            logger.exception("CL handler crashed: %s", e)
        return

    # HubSpot referral channel (refer_a_candidate). HubSpot posts as a bot,
    # so this dispatch must run BEFORE the bot_message subtype filter below.
    if REFER_CANDIDATE_CHANNEL and event.get("channel") == REFER_CANDIDATE_CHANNEL:
        try:
            handle_hubspot_referral_message(event, client)
        except Exception as e:
            logger.exception("HubSpot referral dispatch crashed: %s", e)
        return

    # Only thread replies with files
    if "thread_ts" not in event or "files" not in event:
        return
    if event.get("subtype") in ("bot_message", "message_changed", "message_deleted"):
        return

    thread_ts = event["thread_ts"]
    channel = event["channel"]

    # Get parent message to check if it's an intake confirmation
    try:
        parent = client.conversations_history(
            channel=channel, latest=thread_ts, inclusive=True, limit=1
        )
        if not parent.get("messages"):
            return
        parent_msg = parent["messages"][0]
    except Exception as e:
        logger.debug("Could not fetch parent message: %s", e)
        return

    metadata = parent_msg.get("metadata", {})
    if metadata.get("event_type") == "intake_candidate":
        candidate_id = metadata["event_payload"]["candidate_id"]
        candidate_name = metadata["event_payload"]["candidate_name"]
    else:
        # Fallback: PDF dropped in a reaction-intake thread
        with _reaction_thread_lock:
            entry = _reaction_thread_candidate.get(thread_ts)
        if not entry:
            return
        candidate_id, candidate_name = entry

    # Process PDF files
    for file_info in event.get("files", []):
        mimetype = file_info.get("mimetype", "")
        if mimetype != "application/pdf":
            continue

        file_name = file_info.get("name", "resume.pdf")
        download_url = file_info.get("url_private_download", "")
        if not download_url:
            continue

        try:
            pdf_bytes = _download_slack_file(download_url)
            if not pdf_bytes:
                client.chat_postMessage(
                    channel=channel, thread_ts=thread_ts,
                    text=f":warning: Could not download _{file_name}_ from Slack.",
                )
                continue

            # Upload to Ashby's resume field — this is what the screening
            # pipeline reads (resumeFileHandle → _extract_resume).
            uploaded = _upload_resume_to_ashby(candidate_id, pdf_bytes, file_name)

            if uploaded:
                client.reactions_add(channel=channel, timestamp=event["ts"], name="white_check_mark")
                client.chat_postMessage(
                    channel=channel, thread_ts=thread_ts,
                    text=f":page_facing_up: Resume attached to *{candidate_name}*'s Ashby CV section.",
                )
                logger.info("Resume uploaded for %s from %s (%d bytes)", candidate_name, file_name, len(pdf_bytes))
            else:
                client.chat_postMessage(
                    channel=channel, thread_ts=thread_ts,
                    text=f":x: Failed to upload _{file_name}_ to Ashby. Check logs.",
                )

        except Exception as e:
            logger.error("Failed to process PDF for %s: %s", candidate_name, e)
            client.chat_postMessage(
                channel=channel, thread_ts=thread_ts,
                text=f":x: Failed to process _{file_name}_: {str(e)[:200]}",
            )


# ════════════════════════════════════════════════════════════════
# BULK INTAKE (/intakebulk)
# ════════════════════════════════════════════════════════════════
#
# Flow:
#   1. /intakebulk → initial modal (paste / CSV / XLSX + Event Name + Source)
#   2. Submit → parse → auto-detect roles → push preview modal (paginated)
#   3. Preview: per-row role dropdown + upload button + Next/Back
#   4. Submit preview (Confirm & Push) → batch push to Ashby, dedup, resume upload
#   5. Summary posted in channel
#
# Session state lives in-memory (_BULK_SESSIONS) keyed by a UUID stored in
# each modal's private_metadata.

_BULK_PAGE_SIZE = 20
_BULK_SESSIONS: Dict[str, Dict] = {}  # session_id → session dict


# ── Role keyword buckets for auto-detection ──────────────────────
# Maps a keyword → substring that should appear in a role option's display text.
_ROLE_KEYWORD_BUCKETS: List[Tuple[List[str], List[str]]] = [
    # (keywords that appear in candidate text, substrings to match in role title)
    (["devops", "sre", "site reliability", "platform eng", "infra engineer", "kubernetes", "terraform"], ["devops", "platform", "infra", "sre"]),
    (["frontend", "front-end", "front end", "react", "ui engineer", "web dev"], ["frontend", "front-end", "ui"]),
    (["backend", "back-end", "back end", "api engineer", "server eng"], ["backend", "back-end"]),
    (["full stack", "fullstack", "full-stack"], ["full", "stack"]),
    (["gtm", "go to market", "revenue operations", "revops", "sales ops", "salesforce"], ["gtm", "revenue", "salesforce"]),
    (["product designer", "ux designer", "ui designer", "visual designer"], ["designer", "design"]),
    (["product manager", "technical pm", "tpm", "product management"], ["product", "pm"]),
    (["ml engineer", "machine learning", "ai engineer", "data scientist"], ["ml", "ai", "data"]),
    (["security eng", "appsec", "application security"], ["security"]),
    (["data eng", "analytics eng"], ["data eng", "analytics"]),
]


def _detect_role_option(text: str, role_options: List[Dict]) -> Optional[Dict]:
    """Match candidate text (title/headline/role column) to a role option.

    Returns the best-matching {"text": ..., "value": job_id} or None.
    """
    if not text or not role_options:
        return None
    t = text.lower()

    for keywords, role_substrings in _ROLE_KEYWORD_BUCKETS:
        if any(kw in t for kw in keywords):
            # Find the first role option whose title contains one of the substrings
            for opt in role_options:
                title = opt.get("text", "").lower()
                if any(sub in title for sub in role_substrings):
                    return opt
    return None


# ── Parsers ──────────────────────────────────────────────────────

# CSV column aliases (lowercased). Left = canonical, right = accepted variants.
_COL_ALIASES: Dict[str, List[str]] = {
    "name": ["name", "full name", "candidate", "candidate name"],
    "linkedin": ["linkedin", "linkedin url", "linkedin_url", "li", "profile", "linkedin profile"],
    "email": ["email", "email address", "e-mail"],
    "role": ["role", "target role", "position", "title", "job title", "headline"],
    "notes": ["notes", "note", "comments", "comment"],
    "cv_text": ["cv", "cv_text", "resume", "resume text", "resume_text"],
    "resume_url": ["resume_url", "cv_url", "resume link", "cv link", "pdf url"],
}


def _canonicalize_row(row: Dict) -> Dict:
    """Map a parsed row (arbitrary column names) to canonical keys."""
    out = {"name": "", "linkedin": "", "email": "", "role": "", "notes": "", "cv_text": "", "resume_url": ""}
    low = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items() if k is not None}
    for canon, aliases in _COL_ALIASES.items():
        for alias in aliases:
            if alias in low and low[alias]:
                out[canon] = low[alias]
                break
    return out


def _parse_csv_bytes(data: bytes) -> List[Dict]:
    """Parse CSV bytes into list of canonical row dicts."""
    try:
        text = data.decode("utf-8-sig", errors="replace")
    except Exception:
        text = data.decode("latin-1", errors="replace")
    reader = _csv.DictReader(_io.StringIO(text))
    rows = []
    for raw in reader:
        if not raw:
            continue
        canon = _canonicalize_row(raw)
        if canon["name"] or canon["linkedin"]:
            rows.append(canon)
    return rows


def _parse_xlsx_bytes(data: bytes) -> List[Dict]:
    """Parse XLSX bytes into list of canonical row dicts. Needs openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return []
    try:
        wb = load_workbook(_io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        logger.error("XLSX parse failed: %s", e)
        return []
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return []
    headers = [str(h or "").strip() for h in header]
    rows = []
    for raw_row in it:
        if not raw_row or not any(c is not None and str(c).strip() for c in raw_row):
            continue
        rec = {headers[i]: ("" if raw_row[i] is None else str(raw_row[i])) for i in range(len(headers))}
        canon = _canonicalize_row(rec)
        if canon["name"] or canon["linkedin"]:
            rows.append(canon)
    return rows


_LI_RE = re.compile(r"https?://[^\s|,;]*linkedin\.com/in/[^\s|,;]+", re.I)
_EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")


def _parse_quick_paste(text: str) -> List[Dict]:
    """Parse free-form pasted lines. One candidate per line. Flexible."""
    rows = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        linkedin = ""
        m = _LI_RE.search(line)
        if m:
            linkedin = m.group(0)
            line_wo_li = line.replace(m.group(0), " ")
        else:
            line_wo_li = line

        email = ""
        em = _EMAIL_RE.search(line_wo_li)
        if em:
            email = em.group(0)
            line_wo_li = line_wo_li.replace(em.group(0), " ")

        parts = [p.strip() for p in re.split(r"\s*[|;,\t]\s*", line_wo_li) if p.strip()]
        name = parts[0] if parts else ""
        role_hint = parts[1] if len(parts) > 1 else ""
        notes = " | ".join(parts[2:]) if len(parts) > 2 else ""

        rows.append({
            "name": name, "linkedin": linkedin, "email": email,
            "role": role_hint, "notes": notes, "cv_text": "", "resume_url": "",
        })
    return rows


# ── Session helpers ──────────────────────────────────────────────

def _capture_page_state(session: Dict, view_state: Dict):
    """Read current preview page's state and write selected roles back to session.

    Called before any re-render (Next/Back/Upload/Confirm) so user's role changes stick.
    """
    values = view_state.get("values", {}) if view_state else {}
    page = session.get("page", 0)
    start = page * _BULK_PAGE_SIZE
    rows = session["rows"]
    for i in range(start, min(start + _BULK_PAGE_SIZE, len(rows))):
        block_id = f"brow_{i}"
        block = values.get(block_id, {})
        sel = block.get(f"role_{i}", {}).get("selected_option") or {}
        if sel and sel.get("value"):
            rows[i]["role_job_id"] = sel["value"]
            rows[i]["role_name"] = sel.get("text", {}).get("text", rows[i].get("role_name", ""))


# ── Modal builders ───────────────────────────────────────────────

def build_bulk_initial_modal(channel_id: str) -> dict:
    """First modal: paste + file + event name + source."""
    source_groups = []
    for group in INTAKE_SOURCES:
        opts = [
            {"text": {"type": "plain_text", "text": label}, "value": value}
            for label, value in group["options"]
        ]
        source_groups.append({
            "label": {"type": "plain_text", "text": group["label"]},
            "options": opts,
        })

    # Default to Recruiting Event when uploading files (common use case)
    default_source_initial = {
        "text": {"type": "plain_text", "text": "Sourced: Recruiting Event"},
        "value": "Sourced: Recruiting Event",
    }

    return {
        "type": "modal",
        "callback_id": "bulk_intake_initial",
        "title": {"type": "plain_text", "text": "Bulk Intake"},
        "submit": {"type": "plain_text", "text": "Preview"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": json.dumps({"channel_id": channel_id}),
        "blocks": [
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": (
                        "*Submit multiple candidates at once.*\n"
                        "Paste a block of text, upload a CSV/XLSX, or both.\n"
                        "You'll get a preview to tweak roles + attach resumes before pushing."
                    ),
                },
            },
            {
                "type": "context",
                "elements": [{
                    "type": "mrkdwn",
                    "text": (
                        "CSV/XLSX columns recognized: `name`, `linkedin`, `email`, `role`, "
                        "`notes`, `resume_url`, `cv_text`"
                    ),
                }],
            },
            {"type": "divider"},
            {
                "type": "input",
                "block_id": "event_block",
                "optional": True,
                "label": {"type": "plain_text", "text": "Event Name"},
                "hint": {"type": "plain_text", "text": "E.g. 'AI Summit NYC 2026'. Added to each candidate's note."},
                "element": {
                    "type": "plain_text_input",
                    "action_id": "event_input",
                    "placeholder": {"type": "plain_text", "text": "(optional)"},
                },
            },
            {
                "type": "input",
                "block_id": "source_block",
                "label": {"type": "plain_text", "text": "Default Source (per-row can't be overridden in bulk)"},
                "element": {
                    "type": "static_select",
                    "action_id": "source_select",
                    "option_groups": source_groups,
                    "initial_option": default_source_initial,
                },
            },
            {
                "type": "input",
                "block_id": "paste_block",
                "optional": True,
                "label": {"type": "plain_text", "text": "Paste candidates (one per line)"},
                "hint": {"type": "plain_text", "text": "Format: Name | LinkedIn | Email | Role | Notes. Any order — LinkedIn/email auto-detected."},
                "element": {
                    "type": "plain_text_input",
                    "action_id": "paste_input",
                    "multiline": True,
                    "placeholder": {"type": "plain_text", "text": "Jane Doe | https://linkedin.com/in/jane | jane@x.com | DevOps\nJohn Smith | https://linkedin.com/in/john | Frontend"},
                },
            },
            {
                "type": "input",
                "block_id": "file_block",
                "optional": True,
                "label": {"type": "plain_text", "text": "Upload CSV or XLSX"},
                "hint": {"type": "plain_text", "text": "From a recruiting event? Drop the spreadsheet here."},
                "element": {
                    "type": "file_input",
                    "action_id": "file_input",
                    "filetypes": ["csv", "xlsx", "xls"],
                    "max_files": 1,
                },
            },
        ],
    }


def build_bulk_preview_modal(session_id: str) -> dict:
    """Paginated preview modal. Each row: role dropdown + upload button."""
    session = _BULK_SESSIONS[session_id]
    rows = session["rows"]
    page = session.get("page", 0)
    total = len(rows)
    pages = max(1, (total + _BULK_PAGE_SIZE - 1) // _BULK_PAGE_SIZE)
    start = page * _BULK_PAGE_SIZE
    end = min(start + _BULK_PAGE_SIZE, total)

    role_options = _load_role_options()
    role_opts_blocks = [
        {"text": {"type": "plain_text", "text": r["text"]}, "value": r["value"]}
        for r in role_options
    ]
    role_opts_blocks.append({
        "text": {"type": "plain_text", "text": "No target role / General Pool"},
        "value": OUTBOUND_JOB_ID,
    })

    # Header: stats
    attached = sum(1 for r in rows if r.get("resume_handle") or r.get("resume_bytes_key"))
    header_text = (
        f"*Preview — {total} candidates* | Source: `{session['source']}`"
        + (f" | Event: `{session['event_name']}`" if session.get("event_name") else "")
        + f"\n:page_facing_up: {attached}/{total} have a resume attached."
        + f"\n_Showing {start + 1}–{end} (page {page + 1} of {pages})_"
    )
    blocks: List[Dict] = [
        {"type": "section", "text": {"type": "mrkdwn", "text": header_text}},
        {"type": "divider"},
    ]

    # Rows
    for i in range(start, end):
        row = rows[i]
        name = row.get("name") or "(no name)"
        li = row.get("linkedin") or ""
        role_job_id = row.get("role_job_id") or OUTBOUND_JOB_ID
        role_name = row.get("role_name") or "No target role / General Pool"

        has_resume = bool(row.get("resume_handle") or row.get("resume_bytes_key"))
        status_icon = ":white_check_mark:" if has_resume else ":warning:"
        upload_label = "Replace resume" if has_resume else "Upload resume"

        name_line = f"*{name}*"
        if li:
            name_line += f"  <{li}|LinkedIn>"
        name_line += f"\n{status_icon} " + (
            f"_{row.get('resume_filename', 'resume.pdf')}_" if has_resume else "_No resume_"
        )

        # Find matching initial option for role dropdown
        initial_opt = None
        for opt in role_opts_blocks:
            if opt["value"] == role_job_id:
                initial_opt = opt
                break
        if not initial_opt:
            initial_opt = role_opts_blocks[-1]  # General Pool

        role_select_el = {
            "type": "static_select",
            "action_id": f"role_{i}",
            "options": role_opts_blocks,
            "initial_option": initial_opt,
        }

        blocks.append({
            "type": "section",
            "text": {"type": "mrkdwn", "text": name_line},
        })
        blocks.append({
            "type": "input",
            "block_id": f"brow_{i}",
            "dispatch_action": False,
            "label": {"type": "plain_text", "text": f"Role for {name[:50]}"},
            "element": role_select_el,
        })
        blocks.append({
            "type": "actions",
            "block_id": f"bact_{i}",
            "elements": [{
                "type": "button",
                "action_id": f"upload_{i}",
                "text": {"type": "plain_text", "text": upload_label},
                "value": str(i),
            }],
        })
        blocks.append({"type": "divider"})

    # Navigation (only if multi-page)
    nav_elements = []
    if pages > 1:
        if page > 0:
            nav_elements.append({
                "type": "button",
                "action_id": "bulk_prev",
                "text": {"type": "plain_text", "text": "◀ Previous"},
                "value": "prev",
            })
        if page < pages - 1:
            nav_elements.append({
                "type": "button",
                "action_id": "bulk_next",
                "text": {"type": "plain_text", "text": "Next ▶"},
                "value": "next",
            })
    if nav_elements:
        blocks.append({"type": "actions", "block_id": "nav_block", "elements": nav_elements})

    return {
        "type": "modal",
        "callback_id": "bulk_preview_modal",
        "title": {"type": "plain_text", "text": "Bulk Preview"},
        "submit": {"type": "plain_text", "text": f"Push {total} to Ashby"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": json.dumps({"session_id": session_id}),
        "blocks": blocks,
    }


def build_bulk_row_upload_modal(session_id: str, row_idx: int) -> dict:
    """Small modal to upload a PDF resume for one preview row."""
    session = _BULK_SESSIONS[session_id]
    row = session["rows"][row_idx]
    name = row.get("name") or "(no name)"
    return {
        "type": "modal",
        "callback_id": "bulk_row_upload",
        "title": {"type": "plain_text", "text": "Attach Resume"},
        "submit": {"type": "plain_text", "text": "Attach"},
        "close": {"type": "plain_text", "text": "Cancel"},
        "private_metadata": json.dumps({"session_id": session_id, "row_idx": row_idx}),
        "blocks": [
            {
                "type": "section",
                "text": {"type": "mrkdwn", "text": f"Attach a PDF resume for *{name}*."},
            },
            {
                "type": "input",
                "block_id": "pdf_block",
                "label": {"type": "plain_text", "text": "Resume PDF"},
                "element": {
                    "type": "file_input",
                    "action_id": "pdf_input",
                    "filetypes": ["pdf"],
                    "max_files": 1,
                },
            },
        ],
    }


# ── Bulk handlers ────────────────────────────────────────────────

def handle_bulk_intake(ack, command, client):
    """/intakebulk — open the initial bulk intake modal."""
    ack()
    modal = build_bulk_initial_modal(channel_id=command.get("channel_id", ""))
    client.views_open(trigger_id=command["trigger_id"], view=modal)


def handle_bulk_initial_submission(ack, body, client, view):
    """Parse paste/file, build session, replace modal with preview."""
    values = view["state"]["values"]
    metadata = json.loads(view.get("private_metadata", "{}"))
    channel_id = metadata.get("channel_id", "")
    user_id = body["user"]["id"]

    event_name = (values.get("event_block", {}).get("event_input", {}).get("value") or "").strip()
    source_sel = values["source_block"]["source_select"]["selected_option"]
    source_tag = source_sel["value"]

    paste_text = (values.get("paste_block", {}).get("paste_input", {}).get("value") or "").strip()
    uploaded_files = values.get("file_block", {}).get("file_input", {}).get("files") or []

    rows: List[Dict] = []

    # Parse paste
    if paste_text:
        rows.extend(_parse_quick_paste(paste_text))

    # Parse uploaded file (CSV or XLSX)
    if uploaded_files:
        f = uploaded_files[0]
        download_url = f.get("url_private_download", "")
        file_name = f.get("name", "").lower()
        if download_url:
            data = _download_slack_file(download_url)
            if data:
                if file_name.endswith((".xlsx", ".xls")):
                    rows.extend(_parse_xlsx_bytes(data))
                else:
                    rows.extend(_parse_csv_bytes(data))

    if not rows:
        ack(response_action="errors", errors={
            "paste_block": "No candidates parsed. Paste text or upload a CSV/XLSX."
        })
        return

    # Auto-detect role per row
    role_options = _load_role_options()
    for row in rows:
        hint_text = " ".join([row.get("role", ""), row.get("name", ""), row.get("notes", "")])
        detected = _detect_role_option(hint_text, role_options)
        if detected:
            row["role_job_id"] = detected["value"]
            row["role_name"] = detected["text"]
        else:
            row["role_job_id"] = OUTBOUND_JOB_ID
            row["role_name"] = "No target role / General Pool"
        row["resume_handle"] = ""      # Populated per-row via upload button
        row["resume_filename"] = ""
        row["resume_bytes_key"] = ""   # marker that bytes are stored (pre-upload)

    # Build session
    session_id = str(uuid.uuid4())
    _BULK_SESSIONS[session_id] = {
        "rows": rows,
        "event_name": event_name,
        "source": source_tag,
        "page": 0,
        "user_id": user_id,
        "channel_id": channel_id,
        "resume_bytes": {},  # row_idx → (bytes, filename); uploaded to Ashby on confirm
        "created_at": time.time(),
    }

    # Replace current modal with preview
    preview = build_bulk_preview_modal(session_id)
    ack(response_action="update", view=preview)


def handle_bulk_role_change(ack, body, client):
    """Fires when a role dropdown changes — we just ack; state is captured on nav/confirm."""
    ack()


def handle_bulk_nav(ack, body, client, direction: int):
    """Page navigation. direction: +1 (next) or -1 (prev)."""
    ack()
    view = body["view"]
    metadata = json.loads(view.get("private_metadata", "{}"))
    session_id = metadata.get("session_id", "")
    if session_id not in _BULK_SESSIONS:
        return
    session = _BULK_SESSIONS[session_id]

    # Capture current page's role selections before paging
    _capture_page_state(session, view.get("state", {}))
    session["page"] = max(0, session["page"] + direction)

    updated = build_bulk_preview_modal(session_id)
    client.views_update(view_id=view["id"], view=updated)


def handle_bulk_upload_click(ack, body, client):
    """Open the per-row PDF upload modal on top of the preview."""
    ack()
    action = body["actions"][0]
    view = body["view"]
    metadata = json.loads(view.get("private_metadata", "{}"))
    session_id = metadata.get("session_id", "")
    if session_id not in _BULK_SESSIONS:
        return
    session = _BULK_SESSIONS[session_id]

    # Save current page state before pushing child modal
    _capture_page_state(session, view.get("state", {}))

    row_idx = int(action["value"])
    upload_modal = build_bulk_row_upload_modal(session_id, row_idx)
    client.views_push(trigger_id=body["trigger_id"], view=upload_modal)


def handle_bulk_row_upload_submission(ack, body, client, view):
    """User attached a PDF for a single row. Store bytes in session, refresh preview."""
    metadata = json.loads(view.get("private_metadata", "{}"))
    session_id = metadata.get("session_id", "")
    row_idx = int(metadata.get("row_idx", -1))
    if session_id not in _BULK_SESSIONS or row_idx < 0:
        ack()
        return
    session = _BULK_SESSIONS[session_id]
    rows = session["rows"]
    if row_idx >= len(rows):
        ack()
        return

    pdf_files = view["state"]["values"].get("pdf_block", {}).get("pdf_input", {}).get("files") or []
    if not pdf_files:
        ack(response_action="errors", errors={"pdf_block": "Please upload a PDF."})
        return
    f = pdf_files[0]
    download_url = f.get("url_private_download", "")
    filename = f.get("name", "resume.pdf")
    if not download_url:
        ack(response_action="errors", errors={"pdf_block": "No download URL on uploaded file."})
        return

    pdf_bytes = _download_slack_file(download_url)
    if not pdf_bytes:
        ack(response_action="errors", errors={"pdf_block": "Could not download the PDF from Slack."})
        return

    # Store bytes in session (uploaded to Ashby only on Confirm)
    session["resume_bytes"][row_idx] = (pdf_bytes, filename)
    rows[row_idx]["resume_filename"] = filename
    rows[row_idx]["resume_bytes_key"] = str(row_idx)

    ack()  # close upload modal

    # Refresh the underlying preview modal so the status flips to ✅
    # The root view is under the pushed view in the stack.
    try:
        root_view_id = body.get("view", {}).get("root_view_id") or ""
        if root_view_id:
            updated = build_bulk_preview_modal(session_id)
            client.views_update(view_id=root_view_id, view=updated)
    except Exception as e:
        logger.warning("Could not refresh preview after row upload: %s", e)


def handle_bulk_preview_submission(ack, body, client, view):
    """Confirm & Push — push all rows to Ashby with resumes + dedup."""
    metadata = json.loads(view.get("private_metadata", "{}"))
    session_id = metadata.get("session_id", "")
    if session_id not in _BULK_SESSIONS:
        ack(response_action="errors", errors={"nav_block": "Session expired. Please re-run /intakebulk."})
        return
    session = _BULK_SESSIONS[session_id]

    # Capture final page state
    _capture_page_state(session, view.get("state", {}))

    ack()  # close modal — we report progress in channel

    rows = session["rows"]
    source_tag = session["source"]
    event_name = session.get("event_name", "")
    user_id = session["user_id"]
    channel_id = session["channel_id"]

    # Post initial status
    try:
        client.chat_postEphemeral(
            channel=channel_id, user=user_id,
            text=f":hourglass_flowing_sand: Pushing {len(rows)} candidates to Ashby…",
        )
    except Exception:
        pass

    created: List[Dict] = []
    duplicates: List[Dict] = []
    failed: List[Dict] = []
    resumes_attached = 0

    for idx, row in enumerate(rows):
        name = (row.get("name") or "").strip()
        linkedin = (row.get("linkedin") or "").strip()

        if not name or "linkedin.com/in/" not in linkedin.lower():
            failed.append({"name": name or "(unknown)", "reason": "Missing name or invalid LinkedIn"})
            continue

        # Dedup
        dup = _check_duplicate(name, linkedin)
        if dup:
            duplicates.append({"name": name, "source": dup["source"], "id": dup.get("id", "")})
            continue

        # Build note
        note_parts = [f"<b>Submitted via Slack bulk intake by</b> <@{user_id}>"]
        note_parts.append(f"<b>Source:</b> {source_tag}")
        if event_name:
            note_parts.append(f"<b>Event:</b> {event_name}")
        if row.get("notes"):
            note_parts.append(f"<b>Notes:</b> {row['notes']}")
        if row.get("cv_text"):
            note_parts.append(f"<b>CV / Resume:</b><br>{row['cv_text'][:5000]}")
        notes_html = "<br><br>".join(note_parts)

        # Push
        cid, _aid = _push_to_ashby(
            name=name,
            linkedin=linkedin,
            email=(row.get("email") or "").strip(),
            source=source_tag,
            job_id=row.get("role_job_id") or OUTBOUND_JOB_ID,
            notes_html=notes_html,
        )
        if not cid:
            failed.append({"name": name, "reason": "Ashby create failed"})
            continue

        # Resume: prefer uploaded bytes; otherwise fetch from resume_url if present
        pdf_bytes: Optional[bytes] = None
        filename = "resume.pdf"
        if idx in session["resume_bytes"]:
            pdf_bytes, filename = session["resume_bytes"][idx]
        elif row.get("resume_url"):
            try:
                with urllib.request.urlopen(row["resume_url"], timeout=30) as resp:
                    pdf_bytes = resp.read()
                    filename = os.path.basename(row["resume_url"].split("?")[0]) or "resume.pdf"
            except Exception as e:
                logger.warning("resume_url fetch failed for %s: %s", name, e)

        if pdf_bytes:
            if _upload_resume_to_ashby(cid, pdf_bytes, filename):
                resumes_attached += 1

        created.append({"name": name, "id": cid, "role": row.get("role_name", "")})

    # Cleanup session
    _BULK_SESSIONS.pop(session_id, None)

    # Post summary to channel
    summary_lines = [
        f":white_check_mark: *Bulk intake complete* — submitted by <@{user_id}>",
        f"*Source:* {source_tag}" + (f"  •  *Event:* {event_name}" if event_name else ""),
        "",
        f"• Created: *{len(created)}*",
        f"• Duplicates skipped: *{len(duplicates)}*",
        f"• Failed: *{len(failed)}*",
        f"• Resumes attached: *{resumes_attached}*",
    ]
    if failed:
        summary_lines.append("\n*Failed rows:*")
        for f in failed[:10]:
            summary_lines.append(f"  – {f['name']}: {f['reason']}")
        if len(failed) > 10:
            summary_lines.append(f"  …and {len(failed) - 10} more.")
    if duplicates:
        summary_lines.append("\n*Duplicates skipped:*")
        for d in duplicates[:10]:
            tag = f" (via {d['source']})"
            summary_lines.append(f"  – {d['name']}{tag}")
        if len(duplicates) > 10:
            summary_lines.append(f"  …and {len(duplicates) - 10} more.")

    try:
        client.chat_postMessage(channel=HIRING_CHANNEL, text="\n".join(summary_lines))
    except Exception as e:
        logger.warning("Bulk summary hiring-channel post failed (%s), falling back to ephemeral", e)
        try:
            client.chat_postEphemeral(channel=channel_id, user=user_id, text="\n".join(summary_lines))
        except Exception as e2:
            logger.error("Bulk summary ephemeral fallback failed: %s", e2)

    logger.info(
        "Bulk intake complete: %d created, %d dup, %d failed, %d resumes",
        len(created), len(duplicates), len(failed), resumes_attached,
    )


# ── Screening hooks (Railway-side) ───────────────────────────────

def _format_verdict_summary(res: dict, name: str, candidate_id: str = "", application_id: str = "") -> str:
    verdict = (res.get("verdict") or "").strip() or "SCREENING_FAILED"
    ashby_link = _ashby_candidate_url(candidate_id, application_id) if candidate_id else ""
    link_suffix = f" • <{ashby_link}|Open in Ashby>" if ashby_link else ""

    if verdict == "SCREENING_FAILED":
        err = res.get("error") or "unknown error"
        return f":warning: *{name}* — screening failed.{link_suffix}\n`{err[:300]}`"

    emoji = {
        "SCREEN": ":white_check_mark:",
        "DECLINE": ":x:",
        "DEFER": ":hourglass_flowing_sand:",
        "INSUFFICIENT_DATA": ":grey_question:",
    }.get(verdict, ":page_facing_up:")

    lines = [f"{emoji} *{name}* — screening complete: *{verdict}*{link_suffix}"]
    bf = res.get("best_fit_role") or ""
    if bf:
        lines.append(f"Best fit: _{bf}_")
    return "\n".join(lines)


_INLINE_SCREEN_TIMEOUT_SEC = 600  # 10 min budget for inline screening


def _mark_screening_failed_in_ashby(candidate_id: str, application_id: str,
                                     name: str, reason: str) -> None:
    """Force the candidate into Needs Rescreen with verdict=SCREENING_FAILED.
    Used when the inline screen times out or crashes — gets the candidate onto
    the Needs Rescreen drain so the next ascreen retries them automatically.
    Idempotent: safe to call even if writeback already moved them forward."""
    try:
        import ashby_bridge as ab
        result = {
            "name": name,
            "verdict": "SCREENING_FAILED",
            "verdict_reason": f"Inline screening did not complete: {reason}"[:500],
        }
        ab.write_screening_to_ashby(candidate_id, application_id, result)
    except Exception as e:
        logger.exception("Could not mark %s SCREENING_FAILED in Ashby: %s", name, e)


def _spawn_inline_screen(*, client, candidate_id: str, application_id: str,
                         name: str, thread_ts: str = "",
                         post_channel: str = "",
                         extra_context: str = ""):
    """Run remote_screen.screen_single_candidate in a background thread with a
    watchdog timeout. By default posts to HIRING_CHANNEL (Purple Unicorn);
    pass post_channel to override. extra_context is appended to source_notes
    so Opus sees recruiter-provided dossier (e.g., Candidate Labs write-ups).

    Failure modes:
      - Exception in screening  → post warning, write SCREENING_FAILED to Ashby
      - Watchdog timeout         → post 'will auto-retry', write SCREENING_FAILED
      - Success                  → post verdict + clear pending Slack thread entry

    SCREENING_FAILED routes the candidate to Needs Rescreen automatically (via
    get_verdict_stage). The next ascreen run picks them up and the writeback
    hook posts the verdict back into this same Slack thread."""

    target_channel = post_channel or HIRING_CHANNEL

    # Active-role guard: if the application's job is not on the actively
    # screened list (per ROLE_SHORT_TO_INTERNAL), skip screening with a clear
    # notice. Defense-in-depth — upstream paths (CL, ✅ reaction, /intake)
    # should already filter, but this catches anything that slipped through.
    try:
        from ashby_bridge import _ashby_post as _ab_post
        active_roles = _active_role_names()
        info = _ab_post("application.info", {"applicationId": application_id})
        if info.get("success"):
            job_obj = (info.get("results") or {}).get("job") or {}
            job_title = job_obj.get("title", "") or ""
            # Outbound Sourced is the catch-all for sourced candidates — allow it,
            # post-screening routing will move them to the best-fit active role.
            if job_title and job_title != "Outbound Sourced" and job_title not in active_roles:
                logger.info("Inline screen guard: '%s' is on inactive job '%s' — skipping",
                            name, job_title)
                _safe_post(
                    client,
                    (f":warning: *{name}* — pushed to *{job_title}*, which is not on the "
                     f"active screening list. AI screening skipped; please review manually."),
                    thread_ts, channel=target_channel,
                )
                return
    except Exception as e:
        logger.warning("Active-role guard failed for %s (%s) — proceeding with screen", name, e)

    # Agency + VD Lead exclusion (2026-04-28): bypass AI screening, route to AR.
    # Mirror of the rule in ashby_bridge.pull_for_screening — applied here because
    # inline screening doesn't go through that code path.
    try:
        from ashby_bridge import (
            AGENCY_VD_JOB_TITLES, _ashby_post, add_note, move_to_stage,
            resolve_dest_stage_id,
        )
        info = _ashby_post("application.info", {"applicationId": application_id})
        if info.get("success"):
            app_obj = info.get("results") or {}
            job_title = (app_obj.get("job") or {}).get("title", "")
            current_stage_id = (app_obj.get("currentInterviewStage") or {}).get("id", "")
            cand_obj = app_obj.get("candidate") or {}
            source_obj = cand_obj.get("source") or {}
            source_title = source_obj.get("title", "")
            # Ashby returns source.title as the bare name (e.g. "Quantum") and
            # the category in source.sourceType.title (e.g. "Agencies").
            # Check the category, not a title prefix — see ashby_bridge.py for
            # the bug history (Karan Rami / 2026-04-28).
            source_type_title = ((source_obj.get("sourceType") or {}).get("title") or "").strip().lower()
            if source_type_title == "agencies" and job_title in AGENCY_VD_JOB_TITLES:
                ar_stage_id = resolve_dest_stage_id(current_stage_id, "Application Review")
                if ar_stage_id and current_stage_id != ar_stage_id:
                    move_to_stage(application_id, ar_stage_id)
                add_note(candidate_id, (
                    "<b>Routed to HM Review — AI screening skipped</b><br>"
                    f"Agency-sourced ({source_title}) candidate referred for {job_title}. "
                    "Per policy, agency-referred VD Lead candidates (any level) bypass "
                    "AI screening and go directly to HM review."
                ))
                logger.info("Agency VD Lead — skipping inline screen, routed %s to AR", name)
                _safe_post(
                    client,
                    (f":mag: *{name}* — agency-sourced candidate for *{job_title}*. "
                     f"AI screening skipped per policy; routed to *Application Review* "
                     f"for HM review."),
                    thread_ts, channel=target_channel,
                )
                return
    except Exception as e:
        logger.warning("Agency VD check failed for %s (%s) — proceeding with screen", name, e)

    def _do_screen():
        import remote_screen  # local import: missing deps shouldn't break /intake
        try:
            return remote_screen.screen_single_candidate(
                candidate_id, application_id, extra_context=extra_context
            )
        except TypeError:
            # Older remote_screen signature — retry without extra_context
            return remote_screen.screen_single_candidate(candidate_id, application_id)

    def _worker():
        from concurrent.futures import ThreadPoolExecutor, TimeoutError as FutureTimeout
        try:
            import pending_slack_threads as pst
        except Exception:
            pst = None

        with ThreadPoolExecutor(max_workers=1, thread_name_prefix="inline-screen") as pool:
            future = pool.submit(_do_screen)
            try:
                res = future.result(timeout=_INLINE_SCREEN_TIMEOUT_SEC)
            except FutureTimeout:
                logger.error("Inline screen TIMED OUT after %ds for %s",
                             _INLINE_SCREEN_TIMEOUT_SEC, name)
                _safe_post(
                    client,
                    (f":hourglass_flowing_sand: *{name}* — screening took longer than "
                     f"{_INLINE_SCREEN_TIMEOUT_SEC // 60} minutes; routed to *Needs Rescreen* "
                     f"and will auto-retry on next ascreen run."),
                    thread_ts, channel=target_channel,
                )
                _mark_screening_failed_in_ashby(candidate_id, application_id, name, "watchdog timeout")
                # Leave pending Slack thread entry intact — drain will post verdict.
                return
            except Exception as e:
                logger.exception("Inline screen crashed for %s: %s", name, e)
                _safe_post(
                    client,
                    (f":warning: *{name}* — screening crashed: `{str(e)[:200]}` — "
                     f"routed to *Needs Rescreen* for auto-retry."),
                    thread_ts, channel=target_channel,
                )
                _mark_screening_failed_in_ashby(candidate_id, application_id, name, f"crash: {e}")
                return

        verdict = (res.get("verdict") or "").strip().upper()
        if verdict in ("", "SCREENING_FAILED"):
            err = res.get("error") or "no verdict produced"
            _safe_post(
                client,
                (f":warning: *{name}* — screening did not produce a verdict (`{str(err)[:200]}`); "
                 f"routed to *Needs Rescreen* for auto-retry."),
                thread_ts, channel=target_channel,
            )
            return

        _safe_post(client,
                   _format_verdict_summary(res, name, candidate_id, application_id),
                   thread_ts, channel=target_channel)
        # Clear pending entry — write_screening_to_ashby may have already cleared it
        # via the post-back hook, but pop is idempotent.
        if pst is not None:
            try:
                pst.pop(candidate_id)
            except Exception:
                pass

    t = threading.Thread(target=_worker, name=f"screen-{candidate_id[:8]}", daemon=True)
    t.start()


def _safe_post(client, text: str, thread_ts: str = "", channel: str = ""):
    """Best-effort post to the hiring channel (or an explicit channel)."""
    try:
        client.chat_postMessage(
            channel=channel or HIRING_CHANNEL,
            text=text,
            thread_ts=thread_ts or None,
        )
    except Exception as e:
        logger.warning("Slack post failed: %s", e)


# Approximate API cost per candidate (Opus + Apify + Linkup). Used for the
# confirmation prompt so the user sees the cost impact before approving.
_COST_PER_CANDIDATE_USD = 0.40

# Global screening lock — prevents two /screen runs (or /intake inline screens
# piling on top) from clobbering each other. Holds (running: bool, started_by: str,
# started_at: float). Acquire around the actual batch execution, not the precount.
_screen_lock = threading.Lock()
_screen_state: Dict[str, Any] = {"running": False, "user": "", "started_at": 0.0}


def _screen_busy_message() -> str:
    with _screen_lock:
        if not _screen_state["running"]:
            return ""
        elapsed = int(time.time() - _screen_state["started_at"]) if _screen_state["started_at"] else 0
        who = _screen_state["user"] or "someone"
        return (f":lock: A screening batch is already running (started by <@{who}> "
                f"~{elapsed}s ago). Wait for it to finish before starting another.")


def handle_screen(ack, command, client):
    """`/screen` — count the queue, then ask for confirmation before running."""
    ack(":mag: Counting the screening queue...")
    channel_id = command.get("channel_id", "")
    user_id = command.get("user_id", "")
    response_url = command.get("response_url", "")

    busy = _screen_busy_message()
    if busy:
        _post_to_user(client, channel_id, user_id, text=busy, response_url=response_url)
        return

    def _worker():
        try:
            import remote_screen
            import ashby_bridge as ab
        except Exception as e:
            logger.exception("Screening module import failed: %s", e)
            _safe_post(client, f":warning: Screening module failed to load: `{e}`")
            return

        try:
            dry_pulled = ab.pull_for_screening(dry_run=True, include_leads=True)
        except Exception as e:
            logger.exception("Dry-run pull failed: %s", e)
            _safe_post(client, f":warning: Could not count intake queue: `{e}`")
            return

        intake_count = len(dry_pulled)
        # Rescreen precount is skipped — scanning every active-role Archived app
        # takes too long. Rescreens are discovered + processed during the batch.
        total = intake_count

        if total == 0:
            _post_to_user(client, channel_id, user_id,
                          text=":white_check_mark: Intake queue empty. Any Rescreen-ticked candidates will still be picked up — run /screen anyway if you expect rescreens.",
                          response_url=response_url)
            return

        est_cost = total * _COST_PER_CANDIDATE_USD
        confirm_blocks = [
            {
                "type": "section",
                "text": {
                    "type": "mrkdwn",
                    "text": (
                        f":information_source: *Ready to screen {total} intake candidate(s)*\n"
                        f"• From intake queue (New Lead / App Review / AI Screening): *{intake_count}*\n"
                        f"• Any candidates with the Rescreen checkbox ticked will also be picked up during the run.\n\n"
                        f"Estimated intake API cost: *~${est_cost:.2f}* "
                        f"(_~${_COST_PER_CANDIDATE_USD:.2f}/candidate, rough — rescreens add to this_)"
                    ),
                },
            },
            {
                "type": "actions",
                "elements": [
                    {
                        "type": "button",
                        "style": "primary",
                        "text": {"type": "plain_text", "text": f"Run screening ({total})"},
                        "action_id": "confirm_screen_batch",
                        "value": json.dumps({"user_id": user_id, "total": total}),
                    },
                    {
                        "type": "button",
                        "style": "danger",
                        "text": {"type": "plain_text", "text": "Cancel"},
                        "action_id": "cancel_screen_batch",
                        "value": "cancel",
                    },
                ],
            },
        ]

        _post_to_user(client, channel_id, user_id,
                      text=f"{total} candidates pending — confirm to run.",
                      blocks=confirm_blocks,
                      response_url=response_url)

    threading.Thread(target=_worker, name="screen-precount", daemon=True).start()


def handle_screen_confirm(ack, body, client):
    """User clicked 'Run screening' — kick off the actual batch."""
    ack()
    user_id = body["user"]["id"]
    channel_id = (body.get("channel") or {}).get("id", "")
    response_url = body.get("response_url", "")

    # Guard against a second batch kicking off while one is running
    with _screen_lock:
        if _screen_state["running"]:
            _replace_via_response_url(response_url,
                text=(f":lock: A screening batch is already running (started by "
                      f"<@{_screen_state['user']}>). Try again after it finishes."))
            return
        _screen_state["running"] = True
        _screen_state["user"] = user_id
        _screen_state["started_at"] = time.time()

    # Replace the confirmation message so it can't be clicked again
    _replace_via_response_url(response_url,
                              text=f":hourglass_flowing_sand: Screening batch kicked off by <@{user_id}>.")

    def _worker():
        try:
            import remote_screen
        except Exception as e:
            logger.exception("remote_screen import failed: %s", e)
            _safe_post(client, f":warning: Screening module failed to load: `{e}`")
            with _screen_lock:
                _screen_state["running"] = False
            return

        start = time.time()
        try:
            summary = remote_screen.run_screen_batch()
        except Exception as e:
            logger.exception("/screen batch crashed: %s", e)
            _safe_post(client, f":warning: Screening batch crashed: `{e}`")
            with _screen_lock:
                _screen_state["running"] = False
            return

        elapsed = int(time.time() - start)
        verdicts = summary.get("verdicts") or {}
        verdict_line = ", ".join(f"{v}: {n}" for v, n in sorted(verdicts.items())) or "none"
        summary_text = (
            f":white_check_mark: *Screening batch complete* (triggered by <@{user_id}>)\n"
            f"• Pulled from intake: {summary.get('pulled', 0)}\n"
            f"• Rescreened: {summary.get('rescreened', 0)}\n"
            f"• Verdicts: {verdict_line}\n"
            f"• Errors: {summary.get('errors', 0)}\n"
            f"• Elapsed: {elapsed}s"
        )

        # Per-candidate breakdown: Name | LinkedIn | Verdict
        cands = summary.get("candidates") or []
        if cands:
            order = {"SCREEN": 0, "REVIEW": 1, "DEFER": 2, "INSUFFICIENT_DATA": 3,
                     "DECLINE": 4, "SCREENING_FAILED": 5, "ERROR": 6}
            cands_sorted = sorted(cands, key=lambda c: (order.get(c.get("verdict", ""), 9),
                                                        (c.get("name") or "").lower()))
            lines = []
            for c in cands_sorted:
                nm = c.get("name") or "(no name)"
                li = c.get("linkedin") or "—"
                vd = c.get("verdict") or "?"
                tag = " _(rescreen)_" if c.get("rescreen") else ""
                lines.append(f"• *{vd}* — {nm} — <{li}|LinkedIn>{tag}" if li != "—"
                             else f"• *{vd}* — {nm} — no LinkedIn{tag}")
            summary_text += "\n\n*Candidates:*\n" + "\n".join(lines)

        _safe_post(client, summary_text)

        # Release the global screening lock so a new /screen can be started
        with _screen_lock:
            _screen_state["running"] = False

    threading.Thread(target=_worker, name="screen-batch", daemon=True).start()


def handle_screen_cancel(ack, body, client):
    """User clicked 'Cancel' on the screening confirmation."""
    ack()
    response_url = body.get("response_url", "")
    _replace_via_response_url(response_url,
                              text=":x: Screening batch cancelled. No API calls made.")


def _post_to_user(client, channel_id: str, user_id: str, text: str, blocks=None,
                  response_url: str = ""):
    """Post the message back to the user. Prefers the slash-command response_url
    (works regardless of bot channel membership or scopes), then ephemeral, then DM.
    """
    if response_url:
        try:
            payload: Dict[str, Any] = {"response_type": "ephemeral", "text": text}
            if blocks:
                payload["blocks"] = blocks
            req = urllib.request.Request(
                response_url,
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(req, timeout=10) as resp:
                resp.read()
            return
        except Exception as e:
            logger.warning("response_url post failed (%s), falling back", e)

    try:
        if channel_id:
            client.chat_postEphemeral(
                channel=channel_id, user=user_id, text=text, blocks=blocks,
            )
            return
    except Exception as e:
        logger.warning("Ephemeral post failed (%s), falling back to DM", e)

    try:
        im = client.conversations_open(users=user_id)
        im_channel = im.get("channel", {}).get("id")
        if im_channel:
            client.chat_postMessage(channel=im_channel, text=text, blocks=blocks)
    except Exception as e:
        logger.warning("DM fallback failed: %s", e)


def _replace_via_response_url(response_url: str, text: str):
    """Use Slack's response_url to edit the interactive message."""
    if not response_url:
        return
    try:
        req = urllib.request.Request(
            response_url,
            data=json.dumps({"replace_original": True, "text": text}).encode("utf-8"),
            headers={"Content-Type": "application/json"},
            method="POST",
        )
        urllib.request.urlopen(req, timeout=10).read()
    except Exception as e:
        logger.warning("response_url update failed: %s", e)


# ════════════════════════════════════════════════════════════════
# REACTION-TRIGGERED INTAKE  (#purple-unicorn-hiring)
# ════════════════════════════════════════════════════════════════
#
# React to a message containing LinkedIn URLs with ✅ (white_check_mark) in
# the hiring channel. Bot pulls the URLs, pushes each candidate to Ashby on
# the Outbound Sourced job (New Lead stage), then spawns an inline screen for
# each. Source is inferred from the message + thread + a short prior-message
# window; falls back to "Sourced: LinkedIn".

REACT_TRIGGER_EMOJI = "white_check_mark"
_REACT_SCRAPE_TIMEOUT = 45  # seconds per-URL ceiling for Apify lookup
_REACT_MAX_PARALLEL = 8     # thread pool size for parallel LinkedIn scrapes


def _fetch_real_name(url: str) -> str:
    """Scrape LinkedIn via Apify and return the real display name.
    Returns '' on any failure — caller falls back to slug parsing.
    Result is cached in .linkedin_cache.json so screening reuses it for free."""
    try:
        from apify_linkedin import fetch_linkedin_profile, parse_apify_profile, cache_read, cache_write
    except Exception as e:
        logger.warning("apify_linkedin import failed: %s", e)
        return ""

    token = os.environ.get("APIFY_TOKEN", "")
    if not token:
        return ""

    # Cache-first: free if we've already scraped this URL.
    # cache_read returns the parsed profile dict directly (csv_bridge unwraps
    # the {status, data} envelope) — the old .get("success")/.get("data") check
    # could never pass, so every catch-up cycle re-paid Apify for already-cached
    # people (bug found 2026-06-07).
    cached = cache_read(url)
    if cached and cached.get("name"):
        return (cached.get("name") or "").strip()

    try:
        result = fetch_linkedin_profile(url, token)
    except Exception as e:
        logger.warning("Apify scrape crashed for %s: %s", url, e)
        return ""

    if result.get("success") and result.get("data"):
        cache_write(url, result)
        parsed = parse_apify_profile(result["data"])
        return (parsed.get("name") or "").strip()
    return ""


def _resolve_names_parallel(urls: List[str]) -> Dict[str, str]:
    """Scrape all URLs in parallel. Returns {url: real_name_or_empty}."""
    from concurrent.futures import ThreadPoolExecutor, as_completed
    results: Dict[str, str] = {u: "" for u in urls}
    if not urls:
        return results
    workers = min(_REACT_MAX_PARALLEL, len(urls))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_fetch_real_name, u): u for u in urls}
        for fut in as_completed(futures, timeout=None):
            u = futures[fut]
            try:
                results[u] = fut.result(timeout=_REACT_SCRAPE_TIMEOUT)
            except Exception as e:
                logger.warning("name lookup timed out/failed for %s: %s", u, e)
                results[u] = ""
    return results

_LINKEDIN_URL_RE = re.compile(
    r"https?://(?:[a-z]{2,3}\.)?linkedin\.com/in/[A-Za-z0-9\-_%]+",
    re.I,
)

_processed_reactions: set = set()  # (channel, message_ts) — dedup across multiple reactors
_processed_lock = threading.Lock()

# Maps Slack thread_ts → Ashby candidate_id for reaction-pushed candidates.
# Lets PDFs uploaded AFTER the tick (in the same thread) attach to the right
# candidate. Only populated when exactly one candidate was pushed from a
# reaction — multi-candidate threads are ambiguous for PDF attribution.
_reaction_thread_candidate: Dict[str, Tuple[str, str]] = {}  # thread_ts -> (candidate_id, name)
_reaction_thread_lock = threading.Lock()


def _extract_linkedin_urls(text: str) -> List[str]:
    """Pull unique LinkedIn /in/ URLs from a Slack message body."""
    urls: List[str] = []
    seen: set = set()
    for m in _LINKEDIN_URL_RE.finditer(text or ""):
        url = m.group(0)
        norm = _normalize_linkedin(url)
        if norm and norm not in seen:
            seen.add(norm)
            urls.append(url)
    return urls


def _name_from_linkedin_slug(url: str) -> str:
    """Best-effort display name from a LinkedIn slug.

    'vatsal-shah0914' → 'Vatsal Shah'. This is a placeholder — the screening
    pipeline scrapes LinkedIn and the real name can be set manually in Ashby
    if the guess is off. Returns '' if the slug is unusable."""
    m = re.search(r"/in/([A-Za-z0-9\-_%]+)", url)
    if not m:
        return ""
    slug = m.group(1).replace("%20", "-").replace("_", "-")
    parts: List[str] = []
    for raw in re.split(r"[-]+", slug):
        if not raw or re.fullmatch(r"\d+", raw):
            continue  # pure-digit disambiguator like '279541234'
        stripped = re.sub(r"\d+$", "", raw)  # strip trailing digits: 'shah0914' → 'shah'
        if not stripped or re.search(r"\d", stripped):
            continue  # mixed letter/digit fragment like 'a799611b8' — LinkedIn hash, not a name
        parts.append(stripped.capitalize())
    return " ".join(parts[:4])


# Message-text hints → Ashby source label. Order matters — first match wins.
_SOURCE_HINT_PATTERNS: List[Tuple[Any, str]] = [
    (re.compile(r"\bcandidate\s*labs\b", re.I), "Agencies: Candidate Labs"),
    (re.compile(r"\bhirewell\b", re.I), "Agencies: Hirewell"),
    (re.compile(r"\bdover\b|\bcareers?\s*page\b|\bapplied (through|via|on)\b", re.I), "INBOUND: Dover Careers Page"),
    (re.compile(r"\breferr?al\b|\breferred by\b|\bfriend at\b|\bcolleague at\b|\bmy (friend|colleague)\b", re.I), "Referral: Referral"),
    (re.compile(r"\bjuicebox\b|\bpeoplegpt\b", re.I), "Sourced: Juicebox"),
    (re.compile(r"\bgithub\b", re.I), "Sourced: GitHub"),
    (re.compile(r"\b(conference|hackathon|meetup|recruiting event)\b", re.I), "Sourced: Recruiting Event"),
]


def _detect_source_from_text(text: str) -> str:
    """Infer Ashby source label from context text. Default: Sourced: LinkedIn."""
    for pat, label in _SOURCE_HINT_PATTERNS:
        if pat.search(text or ""):
            return label
    return "Sourced: LinkedIn"


def _gather_thread_pdfs(client, channel: str, message_ts: str, parent_msg: Dict) -> List[Dict]:
    """Collect PDF file objects from:
      - the reacted message itself
      - thread replies (if parent_msg is in a thread)
      - adjacent channel messages ONLY if same-author AND within 5 min
    Returns a list of Slack file dicts (each has name + url_private_download)."""
    pdfs: List[Dict] = []

    def _take_pdfs(msg: Dict) -> None:
        for f in (msg.get("files") or []):
            if (f.get("mimetype") or "") == "application/pdf" and f.get("url_private_download"):
                pdfs.append(f)

    _take_pdfs(parent_msg)

    parent_user = parent_msg.get("user", "")
    try:
        parent_ts_f = float(message_ts)
    except (TypeError, ValueError):
        parent_ts_f = 0.0
    WINDOW_SEC = 300

    t_ts = parent_msg.get("thread_ts") or ""
    try:
        if t_ts:
            thread = client.conversations_replies(channel=channel, ts=t_ts, limit=50).get("messages", [])
            for m in thread:
                if m.get("ts") != message_ts:
                    _take_pdfs(m)
        else:
            prev = client.conversations_history(
                channel=channel, latest=message_ts, limit=6
            ).get("messages", [])
            for m in prev:
                if m.get("ts") == message_ts:
                    continue
                if parent_user and m.get("user", "") != parent_user:
                    continue
                try:
                    m_ts_f = float(m.get("ts", "0"))
                except (TypeError, ValueError):
                    continue
                if parent_ts_f and (parent_ts_f - m_ts_f) > WINDOW_SEC:
                    continue
                _take_pdfs(m)
    except Exception as e:
        logger.debug("reaction-pdf gather failed: %s", e)

    # Dedup by file id
    seen, unique = set(), []
    for f in pdfs:
        fid = f.get("id") or f.get("url_private_download", "")
        if fid and fid not in seen:
            seen.add(fid)
            unique.append(f)
    return unique


def _gather_reaction_context(client, channel: str, message_ts: str, parent_msg: Dict) -> str:
    """Collect text blobs around the reacted message for source detection.

    Only includes:
      - the reacted message itself
      - thread replies on the same message (if any)
      - adjacent channel messages ONLY if by the same author AND within 5 min
        (catches same-user comment batches like 'this is from YC friend' without
         pulling in unrelated prior conversations).
    """
    blobs: List[str] = [parent_msg.get("text", "") or ""]
    parent_user = parent_msg.get("user", "")
    try:
        parent_ts_f = float(message_ts)
    except (TypeError, ValueError):
        parent_ts_f = 0.0
    WINDOW_SEC = 300  # 5 minutes

    t_ts = parent_msg.get("thread_ts") or ""
    try:
        if t_ts:
            thread = client.conversations_replies(channel=channel, ts=t_ts, limit=50).get("messages", [])
            for m in thread:
                if m.get("ts") != message_ts:
                    blobs.append(m.get("text", "") or "")
        else:
            prev = client.conversations_history(
                channel=channel, latest=message_ts, limit=6
            ).get("messages", [])
            for m in prev:
                if m.get("ts") == message_ts:
                    continue
                if parent_user and m.get("user", "") != parent_user:
                    continue
                try:
                    m_ts_f = float(m.get("ts", "0"))
                except (TypeError, ValueError):
                    continue
                if parent_ts_f and (parent_ts_f - m_ts_f) > WINDOW_SEC:
                    continue
                blobs.append(m.get("text", "") or "")
    except Exception as e:
        logger.debug("reaction-context fetch failed: %s", e)

    return "\n".join(blobs)


def handle_reaction(event, client):
    """`reaction_added` → push LinkedIn URLs in the reacted message to Ashby.

    Guardrails:
      - Only the configured trigger emoji (✅) counts.
      - Only fires in the Purple Unicorn Hiring channel.
      - Each message is processed at most once (first reactor wins).
      - If the message has no LinkedIn URLs, silently ignores.
    """
    if event.get("reaction") != REACT_TRIGGER_EMOJI:
        return
    item = event.get("item") or {}
    if item.get("type") != "message":
        return
    channel = item.get("channel", "")
    if channel not in (HIRING_CHANNEL, REFER_CANDIDATE_CHANNEL) or not channel:
        return
    message_ts = item.get("ts", "")
    if not message_ts:
        return

    key = (channel, message_ts)
    with _processed_lock:
        if key in _processed_reactions:
            return
        _processed_reactions.add(key)

    # Backlog path: ✅ on a HubSpot referral message in #refer_a_candidate.
    # Parse the structured payload and dispatch to the same HubSpot pipeline
    # the live message-event handler uses — no LinkedIn-URL inference here.
    if REFER_CANDIDATE_CHANNEL and channel == REFER_CANDIDATE_CHANNEL:
        try:
            resp = client.conversations_history(
                channel=channel, latest=message_ts, inclusive=True, limit=1
            )
            messages = resp.get("messages", []) or []
            if not messages or messages[0].get("ts") != message_ts:
                logger.debug("reacted referral %s not found in history", message_ts)
                return
            parent = messages[0]
        except Exception as e:
            logger.error("Could not fetch referral %s: %s", message_ts, e)
            with _processed_lock:
                _processed_reactions.discard(key)
            return

        text = parent.get("text", "") or ""
        if not text:
            for att in (parent.get("attachments") or []):
                for k in ("text", "fallback", "pretext"):
                    v = att.get(k) or ""
                    if v:
                        text = (text + "\n" + v).strip()

        parsed = _parse_hubspot_referral(text)
        if not parsed:
            logger.info("✅ on %s in refer_a_candidate but no referral payload found", message_ts)
            return

        # Share the live-event dedup set so a back-filled ✅ doesn't double-push
        # if HubSpot's original message_event also fires.
        ref_key = (REFER_CANDIDATE_CHANNEL, message_ts)
        with _processed_referrals_lock:
            if ref_key in _processed_referrals:
                return
            _processed_referrals.add(ref_key)

        try:
            _process_hubspot_referral(client, parent, parsed)
        except Exception as e:
            logger.exception("HubSpot referral backfill crashed: %s", e)
            with _processed_referrals_lock:
                _processed_referrals.discard(ref_key)
        return

    # Fetch the reacted message
    try:
        resp = client.conversations_history(
            channel=channel, latest=message_ts, inclusive=True, limit=1
        )
        messages = resp.get("messages", []) or []
        if not messages or messages[0].get("ts") != message_ts:
            logger.debug("reacted message %s not found in history", message_ts)
            return
        parent = messages[0]
    except Exception as e:
        logger.error("Could not fetch reacted message %s: %s", message_ts, e)
        with _processed_lock:
            _processed_reactions.discard(key)
        return

    msg_text = parent.get("text", "") or ""
    urls = _extract_linkedin_urls(msg_text)
    if not urls:
        # Not an intake message — release the dedup lock so a follow-up edit can retry
        with _processed_lock:
            _processed_reactions.discard(key)
        return

    context_text = _gather_reaction_context(client, channel, message_ts, parent)
    source = _detect_source_from_text(context_text)
    reply_thread_ts = parent.get("thread_ts") or message_ts
    reactor = event.get("user", "")

    # Infer the target role from the reacted message + thread context. If we
    # match exactly one role ("value delivery lead", "backend engineer", etc.)
    # the candidate is created directly on that role's Ashby job instead of
    # landing in the generic Outbound Sourced pool.
    inferred_role, inferred_job_id = _infer_role_from_text(context_text)
    if inferred_role:
        logger.info("Reaction role inference: matched '%s' for message %s",
                    inferred_role, message_ts)

    # Active-role guard: if the message clearly names an inactive role, skip
    # the candidate entirely with a clear warning. Prevents Solution Consultant
    # / Product Analytics Engineer mentions from silently falling back to
    # Outbound Sourced and getting auto-screened.
    if not inferred_role:
        inactive_role = _detect_inactive_role_mention(context_text)
        if inactive_role:
            logger.info("Reaction inactive-role mention '%s' for message %s — skipping",
                        inactive_role, message_ts)
            try:
                client.chat_postMessage(
                    channel=channel,
                    thread_ts=reply_thread_ts,
                    text=(f":warning: *Skipping intake* — message mentioned *{inactive_role}*, "
                          f"which is not on the active screening list. If this candidate "
                          f"should be screened against an active role, please push them manually."),
                )
            except Exception as e:
                logger.warning("inactive-role notice failed: %s", e)
            with _processed_lock:
                _processed_reactions.discard(key)
            return

    target_job_id = inferred_job_id or OUTBOUND_JOB_ID
    role_suffix = f", role: _{inferred_role}_" if inferred_role else ""

    try:
        client.chat_postMessage(
            channel=channel,
            thread_ts=reply_thread_ts,
            text=(f":inbox_tray: <@{reactor}> triggered intake — looking up {len(urls)} "
                  f"LinkedIn profile(s) via Apify, then pushing to Ashby "
                  f"(source: _{source}_{role_suffix}, stage: _New Lead_)..."),
        )
    except Exception as e:
        logger.warning("initial reaction ack failed: %s", e)

    # Parallel LinkedIn lookup for real names. Cache-backed — screening reuses this.
    real_names = _resolve_names_parallel(urls)

    pushed: List[Tuple[str, str, str, str]] = []  # (name, cid, aid, url)
    skipped: List[Tuple[str, str, str]] = []       # (name, url, where)
    errored: List[Tuple[str, str, str]] = []       # (name, url, err)

    for url in urls:
        name = real_names.get(url, "") or _name_from_linkedin_slug(url) or "Unknown (via Slack)"

        dup = _check_duplicate(name, url)
        if dup:
            skipped.append((name, url, dup.get("source", "existing")))
            continue

        notes_html = (
            f"<p>Added via Slack reaction in #purple-unicorn-hiring "
            f"(reacted by &lt;@{html_escape(reactor)}&gt;).</p>"
            f"<p><b>Message:</b> {html_escape(msg_text[:500])}</p>"
        )
        if inferred_role:
            notes_html += f"<p><b>Inferred target role:</b> {html_escape(inferred_role)}</p>"
        try:
            cid, aid = _push_to_ashby(
                name=name,
                linkedin=url,
                email="",
                source=source,
                job_id=target_job_id,
                notes_html=notes_html,
            )
        except Exception as e:
            logger.exception("Slack-reaction push failed for %s: %s", url, e)
            errored.append((name, url, str(e)[:160]))
            continue

        if not cid:
            errored.append((name, url, "candidate.create failed (see logs)"))
            continue

        pushed.append((name, cid, aid or "", url))

    # Summary reply
    summary: List[str] = []
    if pushed:
        summary.append(f":white_check_mark: *Pushed {len(pushed)} to Ashby:*")
        for name, cid, aid, url in pushed:
            link = _ashby_candidate_url(cid, aid)
            summary.append(f"• <{link}|{name}> — {url}")
    if skipped:
        summary.append(f":arrows_counterclockwise: *Skipped {len(skipped)} duplicate(s):*")
        for name, url, where in skipped:
            summary.append(f"• {name} — already in {where} ({url})")
    if errored:
        summary.append(f":x: *Failed {len(errored)}:*")
        for name, url, err in errored:
            summary.append(f"• {name} ({url}) — {err}")

    if summary:
        try:
            client.chat_postMessage(
                channel=channel, thread_ts=reply_thread_ts,
                text="\n".join(summary),
            )
        except Exception as e:
            logger.warning("summary post failed: %s", e)

    # Attach resumes (PDFs found in the reacted message + same-author context).
    # Only safe to auto-attach when exactly one candidate was pushed, otherwise
    # we can't tell which PDF belongs to whom.
    pdf_files = _gather_thread_pdfs(client, channel, message_ts, parent)
    if pdf_files and len(pushed) == 1:
        sole_name, sole_cid, _sole_aid, _sole_url = pushed[0]
        attached = 0
        for f in pdf_files:
            fname = f.get("name", "resume.pdf")
            try:
                pdf_bytes = _download_slack_file(f["url_private_download"])
                if pdf_bytes and _upload_resume_to_ashby(sole_cid, pdf_bytes, fname):
                    attached += 1
                    logger.info("Resume attached to %s from reaction: %s", sole_name, fname)
            except Exception as e:
                logger.warning("reaction-resume upload failed for %s: %s", fname, e)
        if attached:
            try:
                client.chat_postMessage(
                    channel=channel, thread_ts=reply_thread_ts,
                    text=f":page_facing_up: Attached {attached} resume(s) to *{sole_name}*'s Ashby CV section.",
                )
            except Exception as e:
                logger.warning("resume-attach notice failed: %s", e)
    elif pdf_files and len(pushed) > 1:
        try:
            client.chat_postMessage(
                channel=channel, thread_ts=reply_thread_ts,
                text=(f":information_source: Found {len(pdf_files)} PDF(s) in this thread but "
                      f"{len(pushed)} candidates pushed — skipping auto-attach (ambiguous). "
                      f"Reply with a PDF in-thread once you know which candidate it belongs to."),
            )
        except Exception:
            pass

    # Remember thread → candidate mapping so later PDF uploads attach correctly
    if len(pushed) == 1:
        sole_name, sole_cid, _aid, _url = pushed[0]
        with _reaction_thread_lock:
            _reaction_thread_candidate[reply_thread_ts] = (sole_cid, sole_name)

    # Auto-screen each new push. Threads are spawned in parallel; each individual
    # screen internally uses Apify/Linkup/Opus so the wall-clock is the slowest one.
    # context_text (reacted message + thread) goes into Opus as recruiter dossier,
    # same as the Candidate Labs flow passes dossier_text. CV text is extracted by
    # the screening pipeline from the Ashby resume PDF we just uploaded.
    for name, cid, aid, _url in pushed:
        # Register the Slack thread so the verdict post-back can find it even if
        # the candidate gets re-screened later (Needs Rescreen retry).
        try:
            import pending_slack_threads as pst
            pst.register(cid, channel, reply_thread_ts, name=name)
        except Exception as e:
            logger.warning("pending_slack_threads.register failed for %s: %s", name, e)

        _spawn_inline_screen(
            client=client,
            candidate_id=cid,
            application_id=aid,
            name=name,
            thread_ts=reply_thread_ts,
            post_channel=channel,
            extra_context=context_text,
        )


# ════════════════════════════════════════════════════════════════
# CANDIDATE LABS CHANNEL AUTO-INTAKE
# ════════════════════════════════════════════════════════════════
#
# Trigger: a message in CANDIDATE_LABS_CHANNEL from a non-Klarity poster
# (team_id != KLARITY_TEAM_ID) that contains a LinkedIn URL, PLUS a PDF
# that lands in the same thread. The PDF is the "dossier complete" signal.
#
# Flow: silent push to Ashby (source=Agencies: Candidate Labs) → resume
# upload → recruiter dossier (parent + thread text) piped into the Opus
# prompt as extra context → verdict posted to HIRING_CHANNEL (not the
# Candidate Labs channel — external parties read that one).

# parent_ts → {"urls": [...], "dossier_text": "...", "processed": bool}
_cl_pending: Dict[str, Dict[str, Any]] = {}
_cl_pending_lock = threading.Lock()


def _collect_thread_text(client, channel: str, thread_ts: str) -> str:
    """Concatenate the parent message + all thread replies into a single
    plain-text dossier. Used as recruiter-provided context for Opus."""
    try:
        resp = client.conversations_replies(channel=channel, ts=thread_ts, limit=100)
        msgs = resp.get("messages", []) or []
    except Exception as e:
        logger.warning("CL dossier fetch failed for %s: %s", thread_ts, e)
        return ""
    lines: List[str] = []
    for m in msgs:
        txt = (m.get("text") or "").strip()
        if txt:
            lines.append(txt)
    return "\n\n".join(lines)


# ── Role inference from CL Slack context ──────────────────────────
# When a Candidate Labs recruiter posts a candidate and the surrounding
# text mentions a specific role ("value delivery lead", "backend engineer",
# etc.), push the candidate directly onto that role's Ashby job instead of
# the generic Outbound Sourced pool. Falls back to Outbound Sourced when
# no role is mentioned or multiple roles are mentioned ambiguously.
_ROUTING_FILE = _DIR / ".ashby_job_routing.json"
_ROLE_INFERENCE_CACHE: Optional[Dict[str, Any]] = None
_ROLE_INFERENCE_CACHE_LOCK = threading.Lock()


def _active_role_names() -> Set[str]:
    """Roles for which the Opus prompt is configured. Source of truth:
    ROLE_SHORT_TO_INTERNAL in ashby_bridge.py. Slack intake (and inline
    screening) must honor this list — candidates for inactive roles should
    not be pushed or screened."""
    try:
        from ashby_bridge import ROLE_SHORT_TO_INTERNAL
        # Drop "Outbound Sourced" — it's a catch-all job, not a screened role.
        return {v for v in ROLE_SHORT_TO_INTERNAL.values() if v != "Outbound Sourced"}
    except Exception as e:
        logger.warning("Could not load active-role list: %s — defaulting to empty", e)
        return set()


# Keywords identifying the role *families* the Opus prompt is configured to screen.
# Mirrors the active roles in ROLE_SHORT_TO_INTERNAL but matches at the family
# level so Senior/Staff variants of the same role are screened too — the prompt
# itself assesses seniority (junior / senior / staff fit). Anything not matching
# (Solution Consultant, Field Marketing, Alliances Director, Product Marketing
# Manager, Product Analytics Engineer, etc.) gets pushed to Ashby but not inline-screened.
_ACTIVE_ROLE_FAMILY_KEYWORDS = (
    "backend engineer",
    "frontend engineer",
    "product engineer",
    "design engineer",
    "devsecops engineer",
    "gtm engineer",
    "value delivery",
    # Alliances org — REFERRAL PATH ONLY. The dedicated reject-first prompt
    # (opus_body_alliances.md) handles these. Adding them here makes HubSpot/Slack
    # REFERRALS for SC/AD screen inline (verdict as advisory + Taylor's auto-DQ on
    # clean declines + cross-role rescue). Batch `ascreen` is unaffected — it gates
    # on ROLE_SHORT_TO_INTERNAL, so INBOUND SC/AD applicants are never pulled.
    "solution consultant",
    "alliances director",
)


def _is_active_role_family(role_name: str) -> bool:
    """True if `role_name` (e.g. a HubSpot referral role title) belongs to a
    role family the Opus screening prompt is configured for. Family-level match
    so Senior/Staff variants of the 8 base roles screen alongside the base."""
    if not role_name:
        return False
    rn = role_name.lower()
    return any(kw in rn for kw in _ACTIVE_ROLE_FAMILY_KEYWORDS)


def _active_role_job_ids() -> Set[str]:
    """Ashby job IDs for every REFERRAL_ROLE_MAP entry that belongs to an
    active screening role family. Senior / Staff variants of the 8 base roles
    are included (the prompt handles level assessment); truly off-prompt roles
    (Solution Consultant, Alliances Director, etc.) are excluded."""
    try:
        from push_referrals import REFERRAL_ROLE_MAP
    except Exception as e:
        logger.warning("Could not load REFERRAL_ROLE_MAP: %s", e)
        return set()
    return {jid for role_name, jid in REFERRAL_ROLE_MAP.items()
            if jid and _is_active_role_family(role_name)}


def _load_role_inference_patterns() -> Dict[str, Any]:
    """Build {role_name: {"job_id": ..., "patterns": [compiled regex, ...]}}.
    Uses the same .ashby_job_routing.json that post-screening uses, so intake
    routing and best-fit routing stay in sync.

    Filters to ACTIVE roles only (per ROLE_SHORT_TO_INTERNAL). Inactive roles
    like Solution Consultant or Product Analytics Engineer are excluded so
    Slack intake never auto-pushes candidates onto jobs the prompt isn't
    configured for. Detection of inactive-role mentions is handled separately
    by `_detect_inactive_role_mention`."""
    global _ROLE_INFERENCE_CACHE
    with _ROLE_INFERENCE_CACHE_LOCK:
        if _ROLE_INFERENCE_CACHE is not None:
            return _ROLE_INFERENCE_CACHE
        out: Dict[str, Any] = {}
        try:
            data = json.loads(_ROUTING_FILE.read_text(encoding="utf-8"))
        except Exception as e:
            logger.warning("Role inference: could not load %s: %s", _ROUTING_FILE, e)
            _ROLE_INFERENCE_CACHE = {}
            return _ROLE_INFERENCE_CACHE
        active_roles = _active_role_names()
        for role_name, info in (data.get("roles") or {}).items():
            if role_name == "Outbound Sourced":
                continue  # never infer the fallback
            if role_name not in active_roles:
                continue  # inactive role — skip inference entirely
            job_id = info.get("job_id", "")
            if not job_id:
                continue
            phrases = {role_name.lower()}
            for alias in info.get("aliases", []):
                a = (alias or "").strip().lower()
                # Skip short acronyms — too risky for substring false positives
                # (e.g. "VD", "FM", "BE", "FE", "OB", "P+E" inside a URL/token).
                if len(a) <= 3:
                    continue
                phrases.add(a)
            compiled = []
            for p in phrases:
                # Whole-word match, case-insensitive. Treat "+" / "-" as word-ish
                # boundaries rather than regex metacharacters.
                escaped = re.escape(p)
                compiled.append(re.compile(rf"(?<![a-z0-9]){escaped}(?![a-z0-9])", re.I))
            out[role_name] = {"job_id": job_id, "patterns": compiled}
        _ROLE_INFERENCE_CACHE = out
        return _ROLE_INFERENCE_CACHE


_INACTIVE_ROLE_PATTERNS_CACHE: Optional[List[Tuple[str, Any]]] = None
_INACTIVE_ROLE_PATTERNS_LOCK = threading.Lock()


def _load_inactive_role_patterns() -> List[Tuple[str, Any]]:
    """Inactive-role name patterns built from .ashby_job_routing.json minus
    active roles. Used to detect when a Slack message mentions an inactive
    role so we can skip the candidate with a clear warning rather than
    silently default to Outbound Sourced."""
    global _INACTIVE_ROLE_PATTERNS_CACHE
    with _INACTIVE_ROLE_PATTERNS_LOCK:
        if _INACTIVE_ROLE_PATTERNS_CACHE is not None:
            return _INACTIVE_ROLE_PATTERNS_CACHE
        out: List[Tuple[str, Any]] = []
        try:
            data = json.loads(_ROUTING_FILE.read_text(encoding="utf-8"))
        except Exception:
            _INACTIVE_ROLE_PATTERNS_CACHE = []
            return _INACTIVE_ROLE_PATTERNS_CACHE
        active_roles = _active_role_names()
        for role_name, info in (data.get("roles") or {}).items():
            if role_name == "Outbound Sourced":
                continue
            if role_name in active_roles:
                continue  # active — handled by main inference
            phrases = {role_name.lower()}
            for alias in info.get("aliases", []):
                a = (alias or "").strip().lower()
                if len(a) <= 3:
                    continue
                phrases.add(a)
            for p in phrases:
                escaped = re.escape(p)
                out.append((role_name,
                            re.compile(rf"(?<![a-z0-9]){escaped}(?![a-z0-9])", re.I)))
        _INACTIVE_ROLE_PATTERNS_CACHE = out
        return _INACTIVE_ROLE_PATTERNS_CACHE


def _detect_inactive_role_mention(text: str) -> Optional[str]:
    """If `text` clearly names an inactive role (Solution Consultant, Product
    Analytics Engineer, etc.), return that role name. Else None.

    Used by callers to short-circuit intake with a clear warning when the
    recruiter named a role we're not actively screening for. Prevents the
    earlier failure mode where Solution Consultant mentions silently fell
    back to Outbound Sourced and got screened anyway."""
    if not text or not text.strip():
        return None
    for role_name, rx in _load_inactive_role_patterns():
        if rx.search(text):
            return role_name
    return None


def _infer_role_from_text(text: str) -> Tuple[Optional[str], Optional[str]]:
    """Scan free text for a role reference. Returns (role_name, job_id) if
    exactly one distinct role is mentioned, else (None, None).

    Ambiguous matches (multiple distinct roles named) → None, so the caller
    falls back to Outbound Sourced and a human can re-route."""
    if not text or not text.strip():
        return None, None
    patterns = _load_role_inference_patterns()
    if not patterns:
        return None, None
    matched: List[str] = []
    for role_name, info in patterns.items():
        for rx in info["patterns"]:
            if rx.search(text):
                matched.append(role_name)
                break
    distinct = list(dict.fromkeys(matched))  # preserve order, dedup
    if len(distinct) == 1:
        role = distinct[0]
        return role, patterns[role]["job_id"]
    if len(distinct) > 1:
        logger.info("Role inference: ambiguous match (%s) — defaulting to Outbound Sourced",
                    ", ".join(distinct))
    return None, None


def _register_cl_parent(parent_ts: str, urls: List[str]) -> None:
    with _cl_pending_lock:
        if parent_ts not in _cl_pending:
            _cl_pending[parent_ts] = {"urls": urls, "processed": False}


def _claim_cl_parent(parent_ts: str) -> Optional[Dict[str, Any]]:
    """Atomically mark a parent as processed. Returns the entry if the caller
    won the race, else None. Prevents double-push on multiple PDFs in-thread."""
    with _cl_pending_lock:
        entry = _cl_pending.get(parent_ts)
        if not entry or entry.get("processed"):
            return None
        entry["processed"] = True
        return entry


def _post_cl_result(client, header_text: str, thread_ts: str = "") -> Optional[str]:
    """Post to HIRING_CHANNEL. Returns the ts of the new message (so the verdict
    can be threaded under the 'new application' announcement)."""
    try:
        resp = client.chat_postMessage(
            channel=HIRING_CHANNEL,
            text=header_text,
            thread_ts=thread_ts or None,
        )
        return resp.get("ts")
    except Exception as e:
        logger.warning("CL result post failed: %s", e)
        return None


def _process_candidate_labs(client, parent_msg: Dict, pdf_file: Dict, urls: List[str]) -> None:
    """One-shot: push to Ashby + upload resume + screen with recruiter context.
    Verdicts are posted in HIRING_CHANNEL (Purple Unicorn Hiring), never in CL."""
    channel_cl = CANDIDATE_LABS_CHANNEL
    parent_ts = parent_msg.get("ts", "")

    # Dossier = parent text + all thread replies (for Opus context + Ashby note)
    dossier_text = _collect_thread_text(client, channel_cl, parent_ts)

    # Infer the target role from the CL recruiter's message context. If they
    # mention "value delivery lead" / "backend engineer" / etc., push the
    # candidate onto that role's Ashby job directly so it's not sitting in
    # the generic Outbound Sourced pool waiting for post-screening routing.
    inferred_role, inferred_job_id = _infer_role_from_text(dossier_text)
    if inferred_role:
        logger.info("CL role inference: matched '%s' for thread %s",
                    inferred_role, parent_ts)

    # Active-role guard: if the recruiter clearly named a role we're NOT
    # actively screening for (Solution Consultant, Product Analytics Engineer,
    # etc.), skip the candidate entirely with a clear warning. Prevents the
    # earlier failure mode where inactive-role mentions silently fell back to
    # Outbound Sourced and got auto-screened anyway.
    if not inferred_role:
        inactive_role = _detect_inactive_role_mention(dossier_text)
        if inactive_role:
            logger.info("CL inactive-role mention '%s' for thread %s — skipping (silent)",
                        inactive_role, parent_ts)
            return

    target_job_id = inferred_job_id or OUTBOUND_JOB_ID

    # Resolve real name in parallel (cache reused by screening pipeline)
    real_names = _resolve_names_parallel(urls)

    # Download the resume PDF once; upload per candidate (usually 1)
    pdf_bytes = None
    pdf_name = pdf_file.get("name", "resume.pdf")
    try:
        pdf_bytes = _download_slack_file(pdf_file["url_private_download"])
    except Exception as e:
        logger.warning("CL PDF download failed: %s", e)

    for url in urls:
        name = real_names.get(url, "") or _name_from_linkedin_slug(url) or "Unknown (Candidate Labs)"

        dup = _check_duplicate(name, url)
        if dup:
            logger.info(
                "CL dedup skip: %s already in Ashby (cid=%s) — no Slack post",
                name, dup.get("id", "?")[:8],
            )
            continue

        role_line = (
            f"<p><b>Inferred target role:</b> {html_escape(inferred_role)} "
            f"(matched from recruiter context).</p>"
        ) if inferred_role else ""
        notes_html = (
            f"<p><b>Source:</b> {CANDIDATE_LABS_SOURCE} (Slack #candidatelabs-klarity-engineers).</p>"
            f"{role_line}"
            f"<p><b>Recruiter dossier:</b></p><pre>{html_escape(dossier_text[:8000])}</pre>"
        )

        try:
            cid, aid = _push_to_ashby(
                name=name,
                linkedin=url,
                email="",
                source=CANDIDATE_LABS_SOURCE,
                job_id=target_job_id,
                notes_html=notes_html,
            )
        except Exception as e:
            logger.exception("CL push failed for %s: %s", url, e)
            _post_cl_result(
                client,
                f":x: *Candidate Labs push failed* — {name} ({url}) — `{str(e)[:200]}`",
            )
            continue

        if not cid:
            _post_cl_result(
                client,
                f":x: *Candidate Labs push failed* — {name} ({url}) — candidate.create returned no id",
            )
            continue

        # Upload the resume PDF
        if pdf_bytes:
            try:
                _upload_resume_to_ashby(cid, pdf_bytes, pdf_name)
            except Exception as e:
                logger.warning("CL resume upload failed for %s: %s", name, e)

        ashby_link = _ashby_candidate_url(cid, aid or "")
        role_suffix = f" • routed to _{inferred_role}_" if inferred_role else ""
        header_ts = _post_cl_result(
            client,
            (f":inbox_tray: *New application from Candidate Labs* — "
             f"<{ashby_link}|{name}> • source: _{CANDIDATE_LABS_SOURCE}_{role_suffix} • screening started..."),
        )

        # Register the Slack thread so the verdict post-back can find it later —
        # whether the inline screen completes immediately or a Needs Rescreen
        # retry posts it hours later. Persists to .pending_slack_threads.json.
        if header_ts:
            try:
                import pending_slack_threads as pst
                pst.register(cid, HIRING_CHANNEL, header_ts, name=name)
            except Exception as e:
                logger.warning("pending_slack_threads.register failed for %s: %s", name, e)

        _spawn_inline_screen(
            client=client,
            candidate_id=cid,
            application_id=aid or "",
            name=name,
            thread_ts=header_ts or "",
            post_channel=HIRING_CHANNEL,
            extra_context=dossier_text,
        )


def handle_candidate_labs_message(event, client) -> bool:
    """Route messages from the Candidate Labs channel.

    Returns True if the event was handled (so the generic handle_message
    shouldn't reprocess it for /intake-thread PDF logic)."""
    channel = event.get("channel", "")
    if channel != CANDIDATE_LABS_CHANNEL:
        return False
    if event.get("subtype") in ("message_changed", "message_deleted"):
        return True
    # Note: don't filter on team. agency recruiters are
    # guest/Connect users in Klarity's workspace, so their submissions
    # carry the Klarity team ID. The real signal is LinkedIn URL + PDF.

    text = event.get("text", "") or ""
    thread_ts = event.get("thread_ts") or event.get("ts")
    ts = event.get("ts", "")
    urls = _extract_linkedin_urls(text)

    # Case A: this IS the parent (top-level message) with a LinkedIn URL
    if urls and ts == thread_ts:
        _register_cl_parent(thread_ts, urls)
        # If a PDF is already attached to the parent message itself, trigger now
        pdfs = [f for f in (event.get("files") or [])
                if (f.get("mimetype") or "") == "application/pdf"
                and f.get("url_private_download")]
        if pdfs:
            entry = _claim_cl_parent(thread_ts)
            if entry:
                _process_candidate_labs(client, event, pdfs[0], entry["urls"])
        return True

    # Case B: thread reply with a PDF → look up the registered parent
    files = event.get("files") or []
    pdfs = [f for f in files
            if (f.get("mimetype") or "") == "application/pdf"
            and f.get("url_private_download")]
    if not pdfs or not thread_ts:
        return True  # nothing to do, but we handled the event

    with _cl_pending_lock:
        entry = _cl_pending.get(thread_ts)

    if not entry:
        # Parent wasn't captured in-flight (e.g., bot restarted after parent posted).
        # Try to fetch the parent and register it now.
        try:
            hist = client.conversations_history(
                channel=channel, latest=thread_ts, inclusive=True, limit=1
            )
            parent_candidates = hist.get("messages", []) or []
            if parent_candidates and parent_candidates[0].get("ts") == thread_ts:
                parent_text = parent_candidates[0].get("text", "") or ""
                parent_team = parent_candidates[0].get("team", "")
                if parent_team == KLARITY_TEAM_ID:
                    return True
                parent_urls = _extract_linkedin_urls(parent_text)
                if parent_urls:
                    _register_cl_parent(thread_ts, parent_urls)
        except Exception as e:
            logger.debug("CL late-parent lookup failed: %s", e)

    claimed = _claim_cl_parent(thread_ts)
    if not claimed:
        return True

    # Re-fetch the parent message (we need it for _process_candidate_labs)
    try:
        hist = client.conversations_history(
            channel=channel, latest=thread_ts, inclusive=True, limit=1
        )
        parent_msg = (hist.get("messages") or [{}])[0]
    except Exception as e:
        logger.error("CL parent re-fetch failed: %s", e)
        return True

    _process_candidate_labs(client, parent_msg, pdfs[0], claimed["urls"])
    return True


# ════════════════════════════════════════════════════════════════
# HUBSPOT REFERRAL CHANNEL AUTO-INTAKE
# ════════════════════════════════════════════════════════════════
#
# Trigger: HubSpot bot message in REFER_CANDIDATE_CHANNEL with a
# "🎉 New referral" structured payload. We parse the fields, look up
# the role via REFERRAL_ROLE_MAP (push_referrals.py), push to Ashby
# (source = HUBSPOT_REFERRAL_SOURCE, stage = New Lead), then spawn an
# inline screen. Verdict posts to HIRING_CHANNEL — refer_a_candidate
# stays silent (referrers see HubSpot's confirmation, not ours).

# In-process dedup: don't re-process the same message ts if Slack
# redelivers an event during the same boot.
_processed_referrals: set = set()
_processed_referrals_lock = threading.Lock()


def _parse_hubspot_referral(text: str) -> Optional[Dict[str, str]]:
    """Parse a HubSpot 'New referral' Slack message into a flat dict.
    Returns None if the message is not a referral (no anchor + required
    fields missing). Forgiving about whitespace and Slack's <url|label>
    link wrappers."""
    if not text:
        return None
    if "new referral" not in text.lower():
        return None

    # Slack wraps URLs as <url|label> or <url>; emails as <mailto:x|x>.
    # It also HTML-escapes & < > in message text — unescape so role names like
    # "Senior AI Value Delivery & Strategy Lead" match the REFERRAL_ROLE_MAP.
    def _unwrap(s: str) -> str:
        s = re.sub(r"<mailto:([^|>]+)\|[^>]*>", r"\1", s)
        s = re.sub(r"<mailto:([^|>]+)>", r"\1", s)
        s = re.sub(r"<(https?://[^|>]+)\|[^>]*>", r"\1", s)
        s = re.sub(r"<(https?://[^|>]+)>", r"\1", s)
        return html_unescape(s).strip()

    fields: Dict[str, str] = {}
    label_re = re.compile(
        r"^\s*(Referrer Email|Referrer|Candidate|Email|LinkedIn|Role|Strength|Notes)\s*:\s*(.*)$",
        re.IGNORECASE,
    )
    current_key: Optional[str] = None
    for raw_line in text.splitlines():
        line = raw_line.rstrip()
        m = label_re.match(line)
        if m:
            key = m.group(1).strip().lower().replace(" ", "_")
            fields[key] = _unwrap(m.group(2))
            current_key = key
        elif current_key == "notes" and line.strip():
            # Notes can be multi-line — keep appending until next labelled field.
            fields["notes"] = (fields.get("notes", "") + "\n" + _unwrap(line)).strip()

    # Strip trailing "View contact in HubSpot" button-text that some clients
    # render inline at the end of the notes block.
    if fields.get("notes"):
        fields["notes"] = re.sub(
            r"\s*View contact in HubSpot\s*$", "", fields["notes"], flags=re.I
        ).strip()

    # Referrer can come as "Andrew McCann (andrew@klarity.com)" or
    # "Savio Alex Test ()". Split into name + inferred email when paren-form.
    ref = fields.get("referrer", "")
    pm = re.match(r"^\s*(.+?)\s*\(\s*([^\)]*)\s*\)\s*$", ref)
    if pm:
        fields["referrer"] = pm.group(1).strip()
        if pm.group(2).strip() and not fields.get("referrer_email"):
            fields["referrer_email"] = pm.group(2).strip()

    # LinkedIn URL recovery: referrers sometimes fill the form fields wrong —
    # e.g. last name in "LinkedIn" and the actual URL in "Notes" (Beau
    # Davenport, 2026-06-04). If the LinkedIn field holds no linkedin.com URL,
    # scan the whole message for one so dedup keys on the real profile.
    if "linkedin.com" not in fields.get("linkedin", "").lower():
        m_url = re.search(r"https?://(?:www\.)?linkedin\.com/[^\s<>|]+", _unwrap(text), re.I)
        if m_url:
            fields["linkedin"] = m_url.group(0).rstrip(">,.|")

    if not fields.get("candidate") or not fields.get("linkedin"):
        return None
    return fields


def _post_referral_result(client, header_text: str, thread_ts: str = "") -> Optional[str]:
    """Post a referral status line to HIRING_CHANNEL. Returns the ts so
    the inline screen verdict can be threaded under it."""
    try:
        resp = client.chat_postMessage(
            channel=HIRING_CHANNEL,
            text=header_text,
            thread_ts=thread_ts or None,
        )
        return resp.get("ts")
    except Exception as e:
        logger.warning("Referral result post failed: %s", e)
        return None


def _move_to_new_lead(application_id: str) -> None:
    """Move a freshly-created application into the 'New Lead' stage so
    `ascreen` picks it up if the inline screen ever fails. Best-effort —
    a missed move is non-fatal, the verdict-write later overrides anyway.

    Plan-aware: each interview plan has its own 'New Lead' stage with a
    distinct ID. Resolving via the application's current stage keeps the
    move within the correct plan."""
    if not application_id:
        return
    try:
        from ashby_bridge import _ashby_post, resolve_dest_stage_id
        info = _ashby_post("application.info", {"applicationId": application_id})
        current_stage_id = ((info.get("results") or {})
                            .get("currentInterviewStage") or {}).get("id", "")
        stage_id = resolve_dest_stage_id(current_stage_id, "New Lead") if current_stage_id else None
        if not stage_id:
            return
        r = _ashby_post("application.changeStage", {
            "applicationId": application_id,
            "interviewStageId": stage_id,
        })
        if not r.get("success"):
            logger.warning("Referral move→New Lead failed for %s: %s",
                           application_id[:12], str(r)[:300])
    except Exception as e:
        logger.warning("Referral move→New Lead crashed for %s: %s",
                       application_id[:12], e)


def _move_to_referrals_review(application_id: str) -> bool:
    """Move a referral on an OPEN but non-screening-configured role into the
    'Referrals Review' stage (no AI verdict). Returns True if moved.

    Plan-specific: looks up Referrals Review in the application's OWN plan only.
    If that plan has no such stage (e.g. the 'Other referrals' catch-all on the
    Outbound plan, which holds CLOSED roles), returns False and the candidate is
    left in New Lead — we never cross plans into a foreign stage id. Per the recruiting lead,
    2026-05-29."""
    if not application_id:
        return False
    try:
        from ashby_bridge import _ashby_post, load_stages_multi
        info = _ashby_post("application.info", {"applicationId": application_id})
        cur = ((info.get("results") or {}).get("currentInterviewStage") or {})
        current_stage_id = cur.get("id", "")
        plan_id = cur.get("interviewPlanId", "")
        rr_sid = load_stages_multi()["plan_stages"].get(plan_id, {}).get("Referrals Review")
        if not rr_sid or rr_sid == current_stage_id:
            return False
        r = _ashby_post("application.changeStage", {
            "applicationId": application_id,
            "interviewStageId": rr_sid,
        })
        if r.get("success"):
            return True
        logger.warning("Referral move→Referrals Review failed for %s: %s",
                       application_id[:12], str(r)[:300])
        return False
    except Exception as e:
        logger.warning("Referral move→Referrals Review crashed for %s: %s",
                       application_id[:12], e)
        return False


def _resolve_active_application_id(candidate_id: str) -> str:
    """Find a candidate's current non-archived application id. Used as a
    fallback when `application.create` succeeded on Ashby but its response
    timed out on our end (the retry then fails with 'already has active
    application', leaving us without the id). Returns "" if none found."""
    if not candidate_id:
        return ""
    try:
        from ashby_bridge import _ashby_post
        info = _ashby_post("candidate.info", {"id": candidate_id})
        app_ids = ((info.get("results") or {}).get("applicationIds") or [])
        fallback = ""
        for aid in app_ids:
            a = (_ashby_post("application.info", {"applicationId": aid}).get("results") or {})
            fallback = fallback or aid
            if (a.get("status") or "").lower() != "archived":
                return aid
        return fallback
    except Exception as e:
        logger.warning("Could not resolve application id for candidate %s: %s",
                       candidate_id[:12], e)
        return ""


def _archive_self_referral(application_id: str, candidate_id: str = "") -> bool:
    """Archive a self-referral's application (plan-aware), with the
    'Fraudulent Candidate' reason. Returns True if archived. Per the recruiting team
    2026-06-01 — a self-referral is not a genuine referral, so the
    'never auto-decision a referral' safeguard does not apply.

    If `application_id` is missing (e.g. application.create timed out after
    succeeding), falls back to resolving the active application from the
    candidate so a flaky Ashby response can't leave a self-referral unarchived."""
    if not application_id and candidate_id:
        application_id = _resolve_active_application_id(candidate_id)
    if not application_id:
        return False
    try:
        from ashby_bridge import (
            _ashby_post, resolve_dest_stage_id, load_stage_map,
            move_to_stage, SELF_REFERRAL_ARCHIVE_REASON_ID,
        )
        info = _ashby_post("application.info", {"applicationId": application_id})
        current_stage_id = ((info.get("results") or {})
                            .get("currentInterviewStage") or {}).get("id", "")
        arch_id = resolve_dest_stage_id(current_stage_id, "Archived") if current_stage_id else None
        if not arch_id:
            arch_id = load_stage_map().get("Archived")
        if not arch_id:
            logger.error("Self-referral archive failed for %s: no Archived stage resolved",
                         application_id[:12])
            return False
        return move_to_stage(application_id, arch_id, is_archive=True,
                             archive_reason_id=SELF_REFERRAL_ARCHIVE_REASON_ID)
    except Exception as e:
        logger.warning("Self-referral archive crashed for %s: %s", application_id[:12], e)
        return False


def _process_hubspot_referral(client, event: Dict, parsed: Dict[str, str]) -> None:
    """Push parsed referral to Ashby + screen inline. Posts status to
    HIRING_CHANNEL only — refer_a_candidate stays silent."""
    try:
        from push_referrals import REFERRAL_ROLE_MAP, OTHER_JOB_ID
    except Exception as e:
        logger.exception("Referral role map import failed: %s", e)
        REFERRAL_ROLE_MAP = {}
        OTHER_JOB_ID = ""

    name = parsed.get("candidate", "").strip()
    linkedin = parsed.get("linkedin", "").strip()
    email = parsed.get("email", "").strip()
    role = parsed.get("role", "").strip()
    referrer = parsed.get("referrer", "").strip()
    referrer_email = parsed.get("referrer_email", "").strip()
    strength = parsed.get("strength", "").strip()
    notes = parsed.get("notes", "").strip()

    # Self-referral check (recruiting team, 2026-06-01): if the referrer IS the candidate,
    # this isn't a genuine referral — push a record for the audit trail, then
    # auto-archive instead of screening. Decided below, applied after the push.
    from ashby_bridge import is_self_referral
    self_ref = is_self_referral(name, email, referrer, referrer_email)

    # Map role → job_id. Unmapped or closed roles fall into the "Other" job
    # (catch-all parking lot at New Lead, no inline screen).
    job_id = REFERRAL_ROLE_MAP.get(role)
    routed_to_other = False
    if not job_id:
        if not OTHER_JOB_ID:
            _post_referral_result(
                client,
                (f":warning: *HubSpot referral — unmapped role* — *{name}* (`{role}`) "
                 f"from {referrer or 'unknown referrer'}. LinkedIn: {linkedin}. "
                 f"OTHER_JOB_ID not configured."),
            )
            return
        job_id = OTHER_JOB_ID
        routed_to_other = True

    # Dedup against push log + screening log + Ashby name search
    dup = _check_duplicate(name, linkedin)
    if dup:
        logger.info(
            "HubSpot referral dedup skip: %s already in %s (cid=%s) — no Slack post",
            name, dup.get("source", "Ashby"), dup.get("id", "?")[:8],
        )
        return

    # Build a structured note that captures the entire referral payload — the
    # screening pipeline + HMs both read this, so don't truncate the notes.
    referrer_html = html_escape(referrer or "(unknown)")
    if referrer_email:
        referrer_html += f" &lt;{html_escape(referrer_email)}&gt;"
    notes_html_parts = [
        f"<p><b>Source:</b> {HUBSPOT_REFERRAL_SOURCE} (HubSpot referral form, posted in #refer_a_candidate).</p>",
        f"<p><b>Referrer:</b> {referrer_html}</p>",
        f"<p><b>Suggested role:</b> {html_escape(role)}</p>",
    ]
    if strength:
        notes_html_parts.append(f"<p><b>Strength:</b> {html_escape(strength)}</p>")
    if notes:
        notes_html_parts.append(f"<p><b>Referrer notes:</b></p><pre>{html_escape(notes[:6000])}</pre>")
    if self_ref:
        notes_html_parts.insert(
            0,
            "<p>⚠️ <b>SELF-REFERRAL — auto-archived.</b> The referrer and the "
            "candidate are the same person, so this was not treated as a genuine "
            "referral (no screen). Reason: Fraudulent Candidate. Reverse manually "
            "if this was a mistake.</p>",
        )
    notes_html = "".join(notes_html_parts)

    try:
        cid, aid = _push_to_ashby(
            name=name,
            linkedin=linkedin,
            email=email,
            source=HUBSPOT_REFERRAL_SOURCE,
            job_id=job_id,
            notes_html=notes_html,
        )
    except Exception as e:
        logger.exception("HubSpot referral push failed for %s: %s", name, e)
        _post_referral_result(
            client,
            f":x: *HubSpot referral push failed* — {name} ({linkedin}) — `{str(e)[:200]}`",
        )
        return

    if not cid:
        _post_referral_result(
            client,
            f":x: *HubSpot referral push failed* — {name} ({linkedin}) — candidate.create returned no id",
        )
        return

    # Self-referral: archive immediately, skip screening. Still record the
    # referrer fields (documents who self-referred) and flag to the hiring
    # channel so a rare false match can be reversed by hand. (recruiting team, 2026-06-01)
    if self_ref:
        try:
            from ashby_bridge import set_referrer_fields
            if referrer or referrer_email:
                set_referrer_fields(cid, referrer_name=referrer, referrer_email=referrer_email)
        except Exception as e:
            logger.warning("Self-referral referrer-field write failed for %s: %s", name, e)
        archived = _archive_self_referral(aid or "", candidate_id=cid)
        ashby_link = _ashby_candidate_url(cid, aid or "")
        if archived:
            _post_referral_result(
                client,
                (f":no_entry: *Self-referral auto-archived* — <{ashby_link}|{name}> "
                 f"referred themselves (role: _{role or 'unspecified'}_). Not screened. "
                 f"_Reverse manually if this was a mistake._"),
            )
            logger.info("Self-referral auto-archived: %s (cid=%s)", name, cid[:8])
        else:
            _post_referral_result(
                client,
                (f":warning: *Self-referral detected but archive failed* — <{ashby_link}|{name}> "
                 f"referred themselves. Please archive manually."),
            )
            logger.error("Self-referral archive FAILED for %s (cid=%s, aid=%s)",
                         name, cid[:8], (aid or "")[:8])
        return

    # Pin to New Lead so ascreen picks it up if the inline screen fails.
    _move_to_new_lead(aid or "")

    # Credit the internal Klarity referrer (no-op for external referrers).
    try:
        from ashby_bridge import (
            resolve_referrer_user_id, set_application_credited_to,
            set_referrer_fields,
        )
        # Always write Referrer Name + Referrer Email — campaign report needs both.
        if referrer or referrer_email:
            set_referrer_fields(cid, referrer_name=referrer, referrer_email=referrer_email)
        if aid:
            ref_uid = resolve_referrer_user_id(email=referrer_email, name=referrer)
            if ref_uid:
                set_application_credited_to(aid, ref_uid)
    except Exception as e:
        logger.warning("HubSpot referral credit attribution failed for %s: %s", name, e)

    ashby_link = _ashby_candidate_url(cid, aid or "")
    strength_suffix = f" • strength: _{strength.split('—')[0].strip()}_" if strength else ""

    # Gate inline screening on active-role status. Inactive jobs (Solution
    # Consultant, Alliances Director, senior/staff variants etc.) get pushed
    # to Ashby so the referrer's effort isn't lost, but we don't fire the
    # screen because the Opus prompt isn't configured for those roles.
    is_active_role = job_id in _active_role_job_ids()

    # Open-but-unconfigured roles (e.g. Solution Consultant): no screening, but
    # surface the referral in Referrals Review instead of leaving it in New Lead.
    # Closed roles routed to the 'Other' catch-all stay in New Lead — their plan
    # has no Referrals Review stage, so the move is a no-op. Per the recruiting lead 2026-05-29.
    moved_to_rr = False
    if not is_active_role and not routed_to_other:
        moved_to_rr = _move_to_referrals_review(aid or "")

    if is_active_role:
        header_text = (
            f":incoming_envelope: *New HubSpot referral* — <{ashby_link}|{name}> "
            f"from *{referrer or 'unknown'}* • role: _{role}_{strength_suffix} • screening started..."
        )
    elif routed_to_other:
        header_text = (
            f":incoming_envelope: *New HubSpot referral routed to 'Other'* — <{ashby_link}|{name}> "
            f"from *{referrer or 'unknown'}* • requested role: _{role or 'unspecified'}_{strength_suffix} — "
            f"_role not configured / closed, parked for manual review._"
        )
    elif moved_to_rr:
        header_text = (
            f":incoming_envelope: *New HubSpot referral → Referrals Review* — <{ashby_link}|{name}> "
            f"from *{referrer or 'unknown'}* • role: _{role}_{strength_suffix} — "
            f"_role not set up for auto-screening; placed in Referrals Review for manual review._"
        )
    else:
        header_text = (
            f":incoming_envelope: *New HubSpot referral pushed to Ashby* — <{ashby_link}|{name}> "
            f"from *{referrer or 'unknown'}* • role: _{role}_{strength_suffix} — "
            f"_inactive role, no auto-screen._"
        )

    header_ts = _post_referral_result(client, header_text)

    if not is_active_role:
        if routed_to_other:
            logger.info("HubSpot referral %s routed to Other (role='%s' not configured) — skipping inline screen",
                        name, role)
        elif moved_to_rr:
            logger.info("HubSpot referral %s on unconfigured role '%s' → Referrals Review (no screen)",
                        name, role)
        else:
            logger.info("HubSpot referral %s pushed to inactive role '%s' — skipping inline screen",
                        name, role)
        return

    # Persist thread mapping so the verdict post-back works even if the bot
    # restarts mid-screening.
    if header_ts:
        try:
            import pending_slack_threads as pst
            pst.register(cid, HIRING_CHANNEL, header_ts, name=name)
        except Exception as e:
            logger.warning("pending_slack_threads.register failed for %s: %s", name, e)

    # Build dossier text passed to Opus as recruiter context. Mirrors the
    # CL flow — the referrer's notes carry strong signal.
    dossier_lines = [
        f"Referral source: {HUBSPOT_REFERRAL_SOURCE}",
        f"Referrer: {referrer}" + (f" <{referrer_email}>" if referrer_email else ""),
        f"Suggested role: {role}",
    ]
    if strength:
        dossier_lines.append(f"Strength: {strength}")
    if notes:
        dossier_lines.append(f"Referrer notes:\n{notes}")
    dossier_text = "\n".join(dossier_lines)

    _spawn_inline_screen(
        client=client,
        candidate_id=cid,
        application_id=aid or "",
        name=name,
        thread_ts=header_ts or "",
        post_channel=HIRING_CHANNEL,
        extra_context=dossier_text,
    )


def handle_hubspot_referral_message(event, client) -> bool:
    """Route HubSpot bot messages from REFER_CANDIDATE_CHANNEL into the
    referral pipeline. Returns True if the event was claimed."""
    if not REFER_CANDIDATE_CHANNEL:
        return False
    if event.get("channel") != REFER_CANDIDATE_CHANNEL:
        return False
    if event.get("subtype") in ("message_changed", "message_deleted"):
        return True  # claim but ignore edits/deletes

    # HubSpot posts as bot_message. Other subtypes (channel_join, etc.) get
    # claimed but skipped so the generic handler doesn't reprocess them.
    text = event.get("text", "") or ""
    # Bot messages sometimes carry the body in `attachments` rather than `text`.
    if not text:
        for att in (event.get("attachments") or []):
            for k in ("text", "fallback", "pretext"):
                v = att.get(k) or ""
                if v:
                    text = (text + "\n" + v).strip()

    parsed = _parse_hubspot_referral(text)
    if not parsed:
        return True  # not a referral — claim so we don't double-route

    ts = event.get("ts", "")
    key = (REFER_CANDIDATE_CHANNEL, ts)
    with _processed_referrals_lock:
        if key in _processed_referrals:
            return True
        _processed_referrals.add(key)

    try:
        _process_hubspot_referral(client, event, parsed)
    except Exception as e:
        logger.exception("HubSpot referral handler crashed: %s", e)
        with _processed_referrals_lock:
            _processed_referrals.discard(key)  # allow retry on next delivery
    return True


# ── Catch-up scan ────────────────────────────────────────────────
# Slack does not redeliver events that occurred while the bot was offline.
# On startup, replay any candidate-intake messages from the last N hours so
# laptop-off / restart gaps don't silently swallow candidates. Existing dedup
# (Ashby AI Verdict, push log, name+LinkedIn match) makes this safe to re-run.

_CATCH_UP_DEFAULT_HOURS = 48


def _catch_up_candidate_labs(client, oldest: str, latest: str) -> int:
    """Replay any CL parent message (last window) that has a LinkedIn URL +
    a PDF in the thread. Calls the same `_process_candidate_labs` the live
    handler uses — Ashby-side dedup prevents duplicate pushes."""
    if not CANDIDATE_LABS_CHANNEL:
        return 0
    replayed = 0
    cursor = None
    seen = 0
    while True:
        kwargs: Dict[str, Any] = {
            "channel": CANDIDATE_LABS_CHANNEL,
            "oldest": oldest,
            "latest": latest,
            "limit": 100,
        }
        if cursor:
            kwargs["cursor"] = cursor
        try:
            resp = client.conversations_history(**kwargs)
        except Exception as e:
            logger.exception("Catch-up CL: history fetch failed: %s", e)
            return replayed
        messages = resp.get("messages", []) or []
        for msg in messages:
            seen += 1
            if msg.get("subtype") in ("channel_join", "message_changed", "message_deleted"):
                continue
            # Note: don't filter on team. agency recruiters
            # are guest/Connect users in Klarity's workspace, so their
            # submissions carry the Klarity team ID. The real signal is
            # parent + LinkedIn URL + PDF.
            ts = msg.get("ts", "")
            thread_ts = msg.get("thread_ts") or ts
            # Parent messages only (the live handler ignores thread replies as
            # standalone candidates).
            if ts != thread_ts:
                continue
            text = msg.get("text", "") or ""
            urls = _extract_linkedin_urls(text)
            if not urls:
                continue

            # Find a PDF — could be on the parent or in any thread reply.
            pdf_file = None
            for f in (msg.get("files") or []):
                if (f.get("mimetype") or "") == "application/pdf" and f.get("url_private_download"):
                    pdf_file = f
                    break
            if not pdf_file:
                try:
                    replies = client.conversations_replies(
                        channel=CANDIDATE_LABS_CHANNEL, ts=thread_ts, limit=200
                    )
                    for rmsg in replies.get("messages", []) or []:
                        for f in (rmsg.get("files") or []):
                            if (f.get("mimetype") or "") == "application/pdf" and f.get("url_private_download"):
                                pdf_file = f
                                break
                        if pdf_file:
                            break
                except Exception as e:
                    logger.warning("Catch-up CL: replies fetch failed for %s: %s", thread_ts, e)
            if not pdf_file:
                logger.info("Catch-up CL: %s has LinkedIn but no PDF — skipping", thread_ts)
                continue

            logger.info("Catch-up CL: replaying ts=%s urls=%s", thread_ts, urls)
            try:
                _process_candidate_labs(client, msg, pdf_file, urls)
                replayed += 1
            except Exception as e:
                logger.exception("Catch-up CL: replay failed for %s: %s", thread_ts, e)

        cursor = (resp.get("response_metadata") or {}).get("next_cursor") or ""
        if not cursor:
            break

    logger.info("Catch-up CL: scanned %d message(s), replayed %d", seen, replayed)
    return replayed


def _catch_up_hubspot_referrals(client, oldest: str, latest: str) -> int:
    """Replay any HubSpot referral message (last window) in the
    refer_a_candidate channel. Existing dedup prevents duplicate pushes."""
    if not REFER_CANDIDATE_CHANNEL:
        return 0
    replayed = 0
    cursor = None
    seen = 0
    while True:
        kwargs: Dict[str, Any] = {
            "channel": REFER_CANDIDATE_CHANNEL,
            "oldest": oldest,
            "latest": latest,
            "limit": 100,
        }
        if cursor:
            kwargs["cursor"] = cursor
        try:
            resp = client.conversations_history(**kwargs)
        except Exception as e:
            logger.exception("Catch-up HubSpot: history fetch failed: %s", e)
            return replayed
        messages = resp.get("messages", []) or []
        for msg in messages:
            seen += 1
            if msg.get("subtype") in ("message_changed", "message_deleted", "channel_join"):
                continue
            text = msg.get("text", "") or ""
            if not text:
                for att in (msg.get("attachments") or []):
                    for k in ("text", "fallback", "pretext"):
                        v = att.get(k) or ""
                        if v:
                            text = (text + "\n" + v).strip()
            parsed = _parse_hubspot_referral(text)
            if not parsed:
                continue

            ts = msg.get("ts", "")
            key = (REFER_CANDIDATE_CHANNEL, ts)
            with _processed_referrals_lock:
                if key in _processed_referrals:
                    continue
                _processed_referrals.add(key)

            logger.info("Catch-up HubSpot: replaying ts=%s candidate=%s",
                        ts, parsed.get("candidate", "?"))
            try:
                _process_hubspot_referral(client, msg, parsed)
                replayed += 1
            except Exception as e:
                logger.exception("Catch-up HubSpot: replay failed for %s: %s", ts, e)
                with _processed_referrals_lock:
                    _processed_referrals.discard(key)

        cursor = (resp.get("response_metadata") or {}).get("next_cursor") or ""
        if not cursor:
            break

    logger.info("Catch-up HubSpot: scanned %d message(s), replayed %d", seen, replayed)
    return replayed


def _catch_up_scan(client, hours: int = _CATCH_UP_DEFAULT_HOURS) -> None:
    """Top-level catch-up. Bounds the scan window strictly to messages older
    than startup time so we never race with live Socket Mode events."""
    import time as _time
    startup = _time.time()
    oldest = f"{startup - hours * 3600:.6f}"
    latest = f"{startup:.6f}"
    logger.info("Catch-up scan: window = last %dh (oldest=%s, latest=%s)",
                hours, oldest, latest)
    try:
        cl = _catch_up_candidate_labs(client, oldest, latest)
        ref = _catch_up_hubspot_referrals(client, oldest, latest)
        logger.info("Catch-up scan: done — CL=%d, HubSpot=%d", cl, ref)
    except Exception as e:
        logger.exception("Catch-up scan: top-level failure: %s", e)


# ── Main ─────────────────────────────────────────────────────────

def main():
    bot_token = os.environ.get("SLACK_BOT_TOKEN")
    app_token = os.environ.get("SLACK_APP_TOKEN")
    ashby_key = os.environ.get("ASHBY_API_KEY")

    if not bot_token:
        print("\nMissing SLACK_BOT_TOKEN. Set it: export SLACK_BOT_TOKEN='xoxb-...'")
        return
    if not app_token:
        print("\nMissing SLACK_APP_TOKEN. Set it: export SLACK_APP_TOKEN='xapp-...'")
        return
    if not ashby_key:
        print("\nMissing ASHBY_API_KEY.")
        return

    # Initialize Slack app and register handlers
    slack_app = App(token=bot_token)

    # Single intake
    slack_app.command("/intake")(handle_intake)
    slack_app.view("intake_modal")(handle_submission)
    slack_app.event("message")(handle_message)

    # Reaction-triggered intake (✅ on a message with LinkedIn URLs in hiring channel)
    slack_app.event("reaction_added")(handle_reaction)

    # On-demand batch screening (count → confirm → run)
    slack_app.command("/screen")(handle_screen)
    slack_app.action("confirm_screen_batch")(handle_screen_confirm)
    slack_app.action("cancel_screen_batch")(handle_screen_cancel)

    # Bulk intake
    slack_app.command("/intakebulk")(handle_bulk_intake)
    slack_app.view("bulk_intake_initial")(handle_bulk_initial_submission)
    slack_app.view("bulk_preview_modal")(handle_bulk_preview_submission)
    slack_app.view("bulk_row_upload")(handle_bulk_row_upload_submission)
    slack_app.action("bulk_next")(lambda ack, body, client: handle_bulk_nav(ack, body, client, +1))
    slack_app.action("bulk_prev")(lambda ack, body, client: handle_bulk_nav(ack, body, client, -1))
    # Per-row actions (role_N, upload_N) — use regex match
    slack_app.action(re.compile(r"^role_\d+$"))(handle_bulk_role_change)
    slack_app.action(re.compile(r"^upload_\d+$"))(handle_bulk_upload_click)

    logger.info("Slack intake bot starting (Socket Mode)...")
    logger.info("Hiring channel (confirmations): %s", HIRING_CHANNEL)
    logger.info("Candidate Labs channel: %s", CANDIDATE_LABS_CHANNEL)
    logger.info("HubSpot referral channel: %s",
                REFER_CANDIDATE_CHANNEL or "(unset — set REFER_CANDIDATE_CHANNEL_ID env var)")
    logger.info("Roles loaded: %d", len(_load_role_options()))

    # Catch-up scan: replay any candidate-intake messages from the last 48h
    # that the live socket missed (laptop off, restart). Runs in a daemon
    # thread so Socket Mode can start listening immediately. Override the
    # window with CATCH_UP_HOURS env var (e.g. 168 for a week).
    catch_up_hours = int(os.environ.get("CATCH_UP_HOURS", _CATCH_UP_DEFAULT_HOURS))

    def _run_catch_up_thread():
        # Small delay so the socket finishes binding before catch-up posts
        # any verdicts back to threads.
        try:
            time.sleep(3)
        except Exception:
            pass
        _catch_up_scan(slack_app.client, hours=catch_up_hours)

    threading.Thread(
        target=_run_catch_up_thread,
        daemon=True,
        name="catch-up-scan",
    ).start()

    # Periodic catch-up: Socket Mode silently drops events during reconnect
    # gaps (stale connection, WiFi flap, laptop sleep). The startup-only scan
    # misses these because the process never restarts. Re-run the scan every
    # CATCH_UP_INTERVAL_MIN minutes (default 15) so any messages that fell
    # into a gap get replayed within ~15 min. Idempotent — dedup gate skips
    # already-pushed candidates.
    catch_up_interval_min = int(os.environ.get("CATCH_UP_INTERVAL_MIN", 15))

    def _run_periodic_catch_up():
        while True:
            try:
                time.sleep(catch_up_interval_min * 60)
            except Exception:
                pass
            try:
                _catch_up_scan(slack_app.client, hours=catch_up_hours)
            except Exception as e:
                logger.exception("Periodic catch-up failed: %s", e)

    threading.Thread(
        target=_run_periodic_catch_up,
        daemon=True,
        name="catch-up-periodic",
    ).start()

    handler = SocketModeHandler(slack_app, app_token)
    handler.start()


if __name__ == "__main__":
    main()
